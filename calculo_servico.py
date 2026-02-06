from __future__ import annotations

from decimal import Decimal
from typing import Any

import pandas as pd
import streamlit as st

from calculo_servico_graficos import (
    renderizar_grafico_blocos_resumo,
    renderizar_grafico_custos_detalhados,
)


def _converter_para_float(valor: float | int | str | None) -> float | None:
    """Converte valores variados para ``float`` preservando o sinal, quando possível."""

    if valor is None:
        return None

    if isinstance(valor, str):
        normalizado = valor.replace("R$", "").replace("\u00a0", "").strip()

        if not normalizado:
            return None

        sinal = 1
        if normalizado[0] in "+-":
            if normalizado[0] == "-":
                sinal = -1
            normalizado = normalizado[1:]

        normalizado = normalizado.replace(" ", "")

        if not normalizado:
            return None

        if "," in normalizado:
            normalizado = normalizado.replace(".", "").replace(",", ".")
        else:
            if normalizado.count(".") > 1:
                partes = normalizado.split(".")
                if all(parte.isdigit() for parte in partes) and all(
                    len(parte) == 3 for parte in partes[1:]
                ):
                    normalizado = "".join(partes)
                else:
                    normalizado = "".join(partes[:-1]) + "." + partes[-1]
            elif "." in normalizado:
                inteiro, decimal = normalizado.rsplit(".", 1)
                if decimal.isdigit() and len(decimal) == 3 and inteiro.replace(".", "").isdigit():
                    normalizado = inteiro.replace(".", "") + decimal

        if not any(char.isdigit() for char in normalizado):
            return None

        try:
            return sinal * float(normalizado)
        except ValueError:
            return None

    try:
        return float(valor)
    except (TypeError, ValueError):
        return None


def parse_price_to_positive_float(value: Any) -> float:
    """Convert price representations to a positive float value."""

    converted: float | None

    if isinstance(value, (int, float)):
        converted = float(value)
    elif isinstance(value, Decimal):
        converted = float(value)
    elif isinstance(value, str):
        converted = _converter_para_float(value)
    else:
        try:
            converted = float(value)
        except (TypeError, ValueError):
            converted = None

    if converted is None:
        return 0.0

    return abs(converted)


def render_calculo_servico_tab(tab_calculo_servico, format_currency):
    """Renderiza a aba 'Cálculo de serviço'."""
    with tab_calculo_servico:
        st.subheader("🧮 Cálculo de Serviço")

        with st.expander("Cálculo de Instalação", expanded=True):
            total_materiais = parse_price_to_positive_float(
                st.session_state.get("total_custos_materiais", 0.0)
            )
            st.session_state["total_custos_materiais"] = total_materiais
            st.markdown(
                f"<p style='color: black; font-size: 20px;'>🧱 Total Custos com Materiais: {format_currency(total_materiais)}</p>",
                unsafe_allow_html=True,
            )

            total_mao_obra = st.session_state.get("total_custo_mao_obra", 0.0)
            st.markdown(
                f"<p style='color: black; font-size: 20px;'>👷 Total Custo Mão de Obra: {format_currency(total_mao_obra)}</p>",
                unsafe_allow_html=True,
            )

            custo_deslocamento = st.session_state.get("total_custo_deslocamento", 0.0)
            st.markdown(
                f"<p style='color: black; font-size: 20px;'>🚚 Custo estimado de deslocamento: {format_currency(custo_deslocamento)}</p>",
                unsafe_allow_html=True,
            )

            col1, col2 = st.columns([1, 3])
            with col1:
                st.markdown(
                    "<p style='color: black; font-size: 20px;'>➕ Adicional (R$)</p>",
                    unsafe_allow_html=True,
                )
            with col2:
                st.number_input(
                    "Custo adicional",
                    step=0.01,
                    key="custo_adicional",
                    value=st.session_state.get("custo_adicional", 0.0),
                    label_visibility="collapsed",
                )

            custo_adicional = st.session_state.get("custo_adicional", 0.0)

            soma_custos = (
                total_materiais
                + total_mao_obra
                + custo_deslocamento
                + custo_adicional
            )
            st.session_state["soma_parcial_custos"] = soma_custos
            st.markdown(
                f"<p style='color: black; font-size: 20px;'>🧮 Soma parcial dos custos: {format_currency(soma_custos)}</p>",
                unsafe_allow_html=True,
            )

        servicos_adicionais_campos = [
            ("obra_civil", "Obra Civil"),
            ("infra_rede", "Infra de rede"),
            ("andaime", "Andaime"),
            ("transformador", "Transformador"),
            ("totem", "Totem"),
            ("pintura_vaga", "Pintura da vaga"),
            ("pintura_eletrodutos", "Pintura do eletrodutos"),
            ("caminhao_munk", "Caminhão Munk"),
        ]

        servicos_adicionais_selecionados = [
            (chave, nome)
            for chave, nome in servicos_adicionais_campos
            if st.session_state.get(chave) == "Sim"
        ]

        total_servicos_adicionais = 0.0

        if servicos_adicionais_selecionados:
            with st.expander("🔧 Serviços Adicionais", expanded=True):
                for chave, nome in servicos_adicionais_selecionados:
                    campo_custo_base = f"custo_{chave}"
                    campo_custo_calculo = f"calculo_servico_{campo_custo_base}"
                    custo_atual = st.session_state.get(campo_custo_base, 0.0)
                    custo_servico = st.number_input(
                        f"{nome} (R$)",
                        min_value=0.0,
                        step=0.01,
                        key=campo_custo_calculo,
                        value=custo_atual,
                    )
                    st.session_state[campo_custo_base] = custo_servico
                    total_servicos_adicionais += custo_servico

                st.markdown(
                    f"<p style='color: black; font-size: 20px;'>🛠️ Total Serviços Adicionais: {format_currency(total_servicos_adicionais)}</p>",
                    unsafe_allow_html=True,
                )
        else:
            for chave, _ in servicos_adicionais_campos:
                campo_custo_base = f"custo_{chave}"
                campo_custo_calculo = f"calculo_servico_{campo_custo_base}"
                st.session_state[campo_custo_base] = 0.0
                st.session_state[campo_custo_calculo] = 0.0

        st.session_state["total_servicos_adicionais"] = total_servicos_adicionais

        with st.expander("📐 Projeto", expanded=True):
            tipo_servico_atual = (
                st.session_state.get("tipo_servico", "")
                or st.session_state.get("tipo_servico_orcamento", "")
            )
            analise_energia = tipo_servico_atual == "Análise de Energia"
            manutencao_corretiva = tipo_servico_atual == "Manutenção Corretiva"
            projeto_sem_custo = analise_energia or manutencao_corretiva
            if projeto_sem_custo:
                st.session_state["custo_emissao_trt"] = 0.0
                st.session_state["custo_projeto_unifilar"] = 0.0
            col1, col2 = st.columns([1, 3])
            with col1:
                st.markdown(
                    "<p style='color: black; font-size: 20px;'>Emissão de TRT (R$)</p>",
                    unsafe_allow_html=True,
                )
            with col2:
                st.number_input(
                    "Emissão de TRT",
                    min_value=0.0,
                    step=0.01,
                    key="custo_emissao_trt",
                    value=0.0
                    if projeto_sem_custo
                    else st.session_state.get("custo_emissao_trt", 80.0),
                    disabled=projeto_sem_custo,
                    label_visibility="collapsed",
                )

            custo_emissao_trt = 0.0 if projeto_sem_custo else st.session_state.get(
                "custo_emissao_trt", 80.0
            )

            col1, col2 = st.columns([1, 3])
            with col1:
                st.markdown(
                    "<p style='color: black; font-size: 20px;'>Elaboração do Projeto Elétrico (Unifilar) (R$)</p>",
                    unsafe_allow_html=True,
                )
            with col2:
                st.number_input(
                    "Elaboração do Projeto Elétrico (Unifilar)",
                    min_value=0.0,
                    step=0.01,
                    key="custo_projeto_unifilar",
                    value=0.0
                    if projeto_sem_custo
                    else st.session_state.get("custo_projeto_unifilar", 500.0),
                    disabled=projeto_sem_custo,
                    label_visibility="collapsed",
                )

            custo_projeto_unifilar = (
                0.0
                if projeto_sem_custo
                else st.session_state.get("custo_projeto_unifilar", 500.0)
            )

        total_carregadores = 0.0
        st.session_state["preco_unitario_carregador"] = 0.0

        if st.session_state.get("possui_carregador") == "Não":
            dados_carregador = st.session_state.get("carregador_ce_dados", {}) or {}
            preco_carregador = dados_carregador.get("Preço")
            if preco_carregador in (None, ""):
                preco_carregador = st.session_state.get("preco_carregador_orcamento", 0.0)

            preco_carregador_float = parse_price_to_positive_float(preco_carregador)
            st.session_state["preco_unitario_carregador"] = preco_carregador_float

            quantidade_carregadores = st.session_state.get("quantidade_carregadores", 0)
            try:
                quantidade_carregadores_float = float(quantidade_carregadores)
            except (TypeError, ValueError):
                quantidade_carregadores_float = 0.0

            with st.expander("Carregador", expanded=True):
                if dados_carregador:
                    campos_carregador = [
                        ("Fabricante", dados_carregador.get("Fabricante", "")),
                        ("Modelo", dados_carregador.get("Modelo", "")),
                        ("Potência", dados_carregador.get("Potência", "")),
                        ("Tensão", dados_carregador.get("Tensão", "")),
                        ("Carga", dados_carregador.get("Carga", "")),
                        ("Conector", dados_carregador.get("Conector", "")),
                    ]

                    for rotulo, valor in campos_carregador:
                        if valor:
                            st.markdown(f"**{rotulo}:** {valor}")

                    st.markdown(
                        f"**Preço unitário:** {format_currency(preco_carregador_float)}"
                    )

                    col1, col2 = st.columns([1, 3])
                    with col1:
                        st.markdown(
                            "<p style='color: black; font-size: 18px;'>Lucro do carregador (%)</p>",
                            unsafe_allow_html=True,
                        )
                    with col2:
                        st.number_input(
                            "Lucro do carregador (%)",
                            min_value=0.0,
                            step=0.01,
                            key="carregador_lucro_percentual",
                            value=st.session_state.get("carregador_lucro_percentual", 10.0),
                            label_visibility="collapsed",
                        )

                    carregador_lucro_percentual = st.session_state.get(
                        "carregador_lucro_percentual", 10.0
                    )
                    carregador_lucro_valor = (
                        preco_carregador_float * (carregador_lucro_percentual / 100)
                    )

                    col1, col2 = st.columns([1, 3])
                    with col1:
                        st.markdown(
                            "<p style='color: black; font-size: 18px;'>Imposto do carregador (%)</p>",
                            unsafe_allow_html=True,
                        )
                    with col2:
                        st.number_input(
                            "Imposto do carregador (%)",
                            min_value=0.0,
                            step=0.01,
                            key="carregador_imposto_percentual",
                            value=st.session_state.get("carregador_imposto_percentual", 8.0),
                            label_visibility="collapsed",
                        )

                    carregador_imposto_percentual = st.session_state.get(
                        "carregador_imposto_percentual", 8.0
                    )
                    carregador_base_para_imposto = (
                        preco_carregador_float + carregador_lucro_valor
                    )
                    carregador_imposto_valor = (
                        carregador_base_para_imposto
                        * (carregador_imposto_percentual / 100)
                    )

                    preco_carregador_com_acrescimos = (
                        preco_carregador_float
                        + carregador_lucro_valor
                        + carregador_imposto_valor
                    )

                    if quantidade_carregadores_float > 0:
                        lucro_total_carregador = (
                            carregador_lucro_valor * quantidade_carregadores_float
                        )
                        imposto_total_carregador = (
                            carregador_imposto_valor * quantidade_carregadores_float
                        )
                    else:
                        lucro_total_carregador = carregador_lucro_valor
                        imposto_total_carregador = carregador_imposto_valor

                    st.markdown(
                        f"**Lucro total:** {format_currency(lucro_total_carregador)}"
                    )
                    st.markdown(
                        f"**Imposto total:** {format_currency(imposto_total_carregador)}"
                    )

                    st.session_state["preco_unitario_carregador"] = (
                        preco_carregador_com_acrescimos
                    )

                    total_carregadores = preco_carregador_com_acrescimos

                    if quantidade_carregadores:
                        st.markdown(f"**Quantidade:** {int(quantidade_carregadores)}")
                        if preco_carregador_float:
                            total_carregadores = (
                                preco_carregador_com_acrescimos
                                * float(quantidade_carregadores)
                            )
                            st.markdown(
                                f"**Total carregadores:** {format_currency(total_carregadores)}"
                            )
                else:
                    st.info("Nenhum carregador foi selecionado na aba 'Custos com Materiais'.")

        st.session_state["total_carregadores"] = total_carregadores

        base_sem_carregador = (
            total_materiais
            + total_mao_obra
            + custo_deslocamento
            + custo_adicional
            + total_servicos_adicionais
            + custo_emissao_trt
            + custo_projeto_unifilar
        )

        depreciacao = 0.05 * base_sem_carregador
        st.session_state["depreciacao"] = depreciacao

        total_base = base_sem_carregador + depreciacao

        with st.expander("📊 Resultado e Condições", expanded=True):
            st.markdown(
                f"<p style='color: black; font-size: 20px;'>📉 Depreciação: {format_currency(depreciacao)}</p>",
                unsafe_allow_html=True,
            )

            col1, col2 = st.columns([1, 3])
            with col1:
                st.markdown(
                    "<p style='color: black; font-size: 20px;'>Lucro (%)</p>",
                    unsafe_allow_html=True,
                )
            with col2:
                st.number_input(
                    "Lucro (%)",
                    min_value=0.0,
                    step=0.01,
                    key="lucro_percentual",
                    value=st.session_state.get("lucro_percentual", 35.0),
                    label_visibility="collapsed",
                )

            lucro_percentual = st.session_state.get("lucro_percentual", 35.0)
            lucro = (lucro_percentual / 100) * total_base
            st.markdown(
                f"<p style='color: green; font-size: 20px;'>💰 Lucro: {format_currency(lucro)}</p>",
                unsafe_allow_html=True,
            )
            st.session_state["lucro"] = lucro

            # Campo para definir percentual de imposto
            col1, col2 = st.columns([1, 3])
            with col1:
                st.markdown(
                    "<p style='color: black; font-size: 20px;'>Imposto (%)</p>",
                    unsafe_allow_html=True,
                )
            with col2:
                st.number_input(
                    "Imposto (%)",
                    min_value=0.0,
                    step=0.01,
                    key="imposto_percentual",
                    value=st.session_state.get("imposto_percentual", 11.0),
                    label_visibility="collapsed",
                )

            imposto_percentual = st.session_state.get("imposto_percentual", 11.0)
            imposto = (imposto_percentual / 100) * (total_base + lucro)
            st.markdown(
                f"<p style='color: black; font-size: 20px;'>🧾 Imposto: {format_currency(imposto)}</p>",
                unsafe_allow_html=True,
            )
            st.session_state["imposto"] = imposto

        total_projeto = custo_emissao_trt + custo_projeto_unifilar

        total_servico = (
            total_materiais
            + total_mao_obra
            + custo_deslocamento
            + custo_adicional
            + total_servicos_adicionais
            + custo_emissao_trt
            + custo_projeto_unifilar
            + total_carregadores
            + depreciacao
            + lucro
            + imposto
        )
        total_instalacao = total_servico - total_projeto - total_carregadores

        total_instalacao_formatado = format_currency(total_instalacao)
        st.session_state["total_instalacao_calculo_servico"] = total_instalacao_formatado
        st.session_state["total_instalacao_valor"] = total_instalacao

        st.markdown(
            f"<p style='color: black; font-size: 25px;'>🏗️ Total Instalação: {total_instalacao_formatado}</p>",
            unsafe_allow_html=True,
        )
        st.markdown(
            f"<p style='color: black; font-size: 30px;'>💰 Total Serviço: {format_currency(total_servico)}</p>",
            unsafe_allow_html=True,
        )
        st.session_state["total_calculo_servico"] = total_servico

        st.markdown("---")
        st.subheader("📈 Visualização dos Custos")

        valor_projeto = custo_emissao_trt + custo_projeto_unifilar
        valor_calculo_instalacao = (
            total_materiais + total_mao_obra + custo_deslocamento + custo_adicional
        )

        renderizar_grafico_custos_detalhados(
            dados={
                "materiais": total_materiais,
                "mao_obra": total_mao_obra,
                "deslocamento": custo_deslocamento,
                "adicional": custo_adicional,
                "servicos_adicionais": total_servicos_adicionais,
                "projeto": valor_projeto,
                "carregadores": total_carregadores,
                "depreciacao": depreciacao,
                "lucro": lucro,
                "imposto": imposto,
            }
        )

        renderizar_grafico_blocos_resumo(
            dados={
                "calculo_instalacao": valor_calculo_instalacao,
                "servicos_adicionais": total_servicos_adicionais,
                "projeto": valor_projeto,
                "depreciacao": depreciacao,
                "lucro": lucro,
                "imposto": imposto,
            }
        )

        if st.button("Salvar Dados do Cálculo de Serviço"):
            try:
                from openpyxl import load_workbook
            except ImportError:
                load_workbook = None

            from pathlib import Path

            docs_dir = Path(__file__).with_name("Docs Salvos")
            docs_dir.mkdir(exist_ok=True)

            ordem_venda = st.session_state.get("ordem_venda", "")
            base_name = "Dados do Cálculo de Serviço"
            filename = (
                f"{base_name} {ordem_venda}.xlsx"
                if ordem_venda
                else f"{base_name}.xlsx"
            )
            filepath = docs_dir / filename

            dados_calculo = {
                "Ordem de Venda": ordem_venda,
                "Cliente": st.session_state.get("cliente", ""),
                "Tipo de Serviço": st.session_state.get("tipo_servico", ""),
                "Total Materiais": float(total_materiais),
                "Total Mão de Obra": float(total_mao_obra),
                "Custo Deslocamento": float(custo_deslocamento),
                "Custo Adicional": float(custo_adicional),
                "Total Serviços Adicionais": float(total_servicos_adicionais),
                "Custo Emissão TRT": float(custo_emissao_trt),
                "Custo Projeto Unifilar": float(custo_projeto_unifilar),
                "Total Projeto": float(total_projeto),
                "Total Carregadores": float(total_carregadores),
                "Depreciação": float(depreciacao),
                "Lucro (%)": float(lucro_percentual),
                "Lucro": float(lucro),
                "Imposto (%)": float(imposto_percentual),
                "Imposto": float(imposto),
                "Soma Parcial Custos": float(soma_custos),
                "Base Sem Carregador": float(base_sem_carregador),
                "Base Total": float(total_base),
                "Total Instalação": float(total_instalacao),
                "Total Serviço": float(total_servico),
            }

            df = pd.DataFrame([dados_calculo])

            if filepath.exists():
                if load_workbook is None:
                    st.error(
                        "Não foi possível atualizar o arquivo Excel porque a biblioteca "
                        "'openpyxl' não está instalada. Execute `pip install openpyxl` e tente novamente."
                    )
                else:
                    book = load_workbook(filepath)
                    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        startrow = book.active.max_row
                        df.to_excel(writer, index=False, header=False, startrow=startrow)
                    st.success(f"Dados salvos em {filepath}")
            else:
                try:
                    df.to_excel(filepath, index=False)
                except ImportError:
                    st.error(
                        "Não foi possível criar o arquivo Excel porque a biblioteca "
                        "'openpyxl' não está instalada. Execute `pip install openpyxl` e tente novamente."
                    )
                else:
                    st.success(f"Dados salvos em {filepath}")
