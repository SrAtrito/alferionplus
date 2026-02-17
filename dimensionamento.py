import streamlit as st
import math
import re
from pathlib import Path
from collections.abc import Mapping
from typing import Any, Optional
import pandas as pd
from dados_transformadores import obter_produtos_transformadores_padrao
from tabelas_eletricas import (
    TABELA_BITOLAS,
    TABELA_NEUTRO_TERRA,
    TABELA_CABO_ISOLADO_PVC,
    TABELA_CABO_UNIPOLAR_HEPR,
    TABELA_ELETRODUTOS,
)


def _sincronizar_tamanho_eletroduto_personalizado():
    """Mantém o tamanho de eletroduto personalizado sincronizado com o estado."""
    valor_personalizado = st.session_state.get("tamanho_eletroduto_resumo", "")
    st.session_state["tamanho_eletroduto"] = valor_personalizado
    sugerido = st.session_state.get("tamanho_eletroduto_sugerido", "")
    st.session_state["tamanho_eletroduto_manual"] = valor_personalizado != sugerido


def _load_editor_dataframe(state_key: str) -> Optional[pd.DataFrame]:
    """Carrega um DataFrame salvo em session_state pelos editores de tabela."""
    raw = st.session_state.get(state_key)
    if raw is None:
        return None
    if isinstance(raw, pd.DataFrame):
        return raw
    if isinstance(raw, list):
        return pd.DataFrame(raw)
    if isinstance(raw, Mapping):
        if "data" in raw and "columns" in raw:
            return pd.DataFrame(raw["data"], columns=raw["columns"])
        return pd.DataFrame(raw)
    return pd.DataFrame(raw)


def _obter_opcoes_material(state_key: str, arquivo_csv: str) -> list[str]:
    """Retorna opções de materiais a partir do editor ou do arquivo CSV."""
    df = _load_editor_dataframe(state_key)
    if df is None or df.empty:
        try:
            df = pd.read_csv(arquivo_csv, sep=";")
        except FileNotFoundError:
            return [""]
    if "Material" not in df.columns:
        return [""]
    opcoes: list[str] = [""]
    vistos: set[str] = set()
    for item in df["Material"]:
        if pd.isna(item):
            continue
        texto = str(item).strip()
        if not texto or texto.lower() == "nan":
            continue
        if texto not in vistos:
            vistos.add(texto)
            opcoes.append(texto)
    return opcoes


def _render_selectbox_material(
    label: str, key: str, opcoes: list[str], sugestao: str
) -> None:
    """Renderiza um selectbox garantindo que opções atuais e sugeridas estejam disponíveis."""
    if not opcoes:
        opcoes_norm = [""]
    else:
        opcoes_norm = list(opcoes)
    if opcoes_norm[0] != "":
        opcoes_norm = [""] + [opt for opt in opcoes_norm if opt]
    if sugestao and sugestao not in opcoes_norm:
        opcoes_norm.append(sugestao)
    valor_atual = st.session_state.get(key, "")
    if valor_atual and valor_atual not in opcoes_norm:
        opcoes_norm.append(valor_atual)
    indice = opcoes_norm.index(valor_atual) if valor_atual in opcoes_norm else 0
    st.selectbox(
        label,
        options=opcoes_norm,
        index=indice,
        key=key,
        label_visibility="collapsed",
    )


def _atualizar_corrente_nominal_manual() -> None:
    """Atualiza o estado que indica se a corrente nominal foi ajustada manualmente."""
    corrente_calculada = st.session_state.get("corrente_calculada", "")
    corrente_nominal = st.session_state.get("corrente_nominal", "")
    st.session_state["corrente_nominal_manual"] = corrente_nominal != corrente_calculada


def _normalizar_valor_bitola(valor) -> Optional[float]:
    """Converte representações de bitola em um valor numérico."""

    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        try:
            return float(valor)
        except (TypeError, ValueError):
            return None
    texto = str(valor).strip()
    if not texto:
        return None
    texto = (
        texto.replace("mm²", "")
        .replace("mm", "")
        .replace(" ", "")
        .replace(",", ".")
    )
    if not texto:
        return None
    try:
        return float(texto)
    except ValueError:
        return None


def _formatar_bitola_mm2(valor) -> str:
    """Formata um valor numérico de bitola para exibição."""

    numero = _normalizar_valor_bitola(valor)
    if numero is None:
        return ""
    if numero.is_integer():
        texto = str(int(numero))
    else:
        texto = f"{numero:.1f}".rstrip("0").rstrip(".")
        texto = texto.replace(".", ",")
    return f"{texto} mm²"


def _obter_opcoes_bitola_cabos(tabela: pd.DataFrame) -> list[str]:
    """Gera as opções de bitola disponíveis a partir da tabela selecionada."""

    if "Cabo (mm²)" not in tabela.columns:
        return [""]
    valores = {
        numero
        for numero in (
            _normalizar_valor_bitola(item) for item in tabela["Cabo (mm²)"]
        )
        if numero is not None
    }
    opcoes = [""] + [
        _formatar_bitola_mm2(valor) for valor in sorted(valores)
    ]
    return opcoes


def _obter_configuracao_tabela_cabos(tipo_cabos: str) -> tuple[str, pd.DataFrame, str]:
    """Retorna as chaves e tabela base associadas ao tipo de cabo selecionado."""

    if tipo_cabos == "Cabo PVC":
        return "tabela_cabos_pvc", TABELA_CABO_ISOLADO_PVC, "tabela_cabos_editor_pvc"
    return "tabela_cabos_hepr", TABELA_CABO_UNIPOLAR_HEPR, "tabela_cabos_editor_hepr"


def _normalizar_tabela_cabos(
    tabela: Any, tabela_base: pd.DataFrame
) -> pd.DataFrame:
    """Garante que a tabela de cabos esteja no formato DataFrame esperado."""

    if not isinstance(tabela, pd.DataFrame):
        if isinstance(tabela, dict):
            tabela = {
                chave: list(valor.values()) if isinstance(valor, dict) else valor
                for chave, valor in tabela.items()
            }
        try:
            tabela = pd.DataFrame(tabela)
            if "Cabo (mm²)" not in tabela.columns:
                tabela.columns = tabela_base.columns[: len(tabela.columns)]
            tabela = tabela.reindex(columns=tabela_base.columns)
        except (ValueError, TypeError):
            tabela = tabela_base.copy()
    else:
        tabela = tabela.copy()

    if tabela.empty:
        tabela = tabela_base.copy()
    return tabela


def _registrar_bitola_calculada(
    chave_valor: str, chave_sugestao: str, chave_manual: str, valor: str
) -> None:
    """Armazena a bitola calculada e respeita ajustes manuais existentes."""

    valor_normalizado = valor or ""
    sugestao_anterior = st.session_state.get(chave_sugestao, valor_normalizado)
    manual_atual = st.session_state.get(chave_manual, False)

    if (
        not manual_atual
        and chave_valor in st.session_state
        and st.session_state.get(chave_valor) != sugestao_anterior
    ):
        manual_atual = True
        st.session_state[chave_manual] = True

    if chave_manual not in st.session_state:
        st.session_state[chave_manual] = manual_atual
    if chave_valor not in st.session_state:
        st.session_state[chave_valor] = valor_normalizado

    st.session_state[chave_sugestao] = valor_normalizado

    if (
        not st.session_state.get(chave_manual, False)
        or not st.session_state.get(chave_valor)
    ):
        st.session_state[chave_valor] = valor_normalizado


def _render_bitola_select(
    label: str,
    chave_valor: str,
    chave_sugestao: str,
    chave_manual: str,
    opcoes: list[str],
    cor: Optional[str] = None,
) -> None:
    """Exibe um seletor de bitola permitindo ajustes manuais."""

    sugestao = st.session_state.get(chave_sugestao, "")
    valor_atual = st.session_state.get(chave_valor, sugestao)
    if valor_atual is None:
        valor_atual = ""
    st.session_state.setdefault(chave_manual, False)
    opcoes_select = list(opcoes)
    if sugestao and sugestao not in opcoes_select:
        opcoes_select.append(sugestao)
    if valor_atual and valor_atual not in opcoes_select:
        opcoes_select.append(valor_atual)
    # Garante que a opção vazia esteja presente e na primeira posição.
    if "" not in opcoes_select:
        opcoes_select.insert(0, "")
    else:
        opcoes_select = [""] + [opt for opt in opcoes_select if opt != ""]

    if cor:
        st.markdown(
            f"<p style='font-size:24px; font-weight:bold; color:{cor};'>{label}</p>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"<p style='font-size:24px; font-weight:bold;'>{label}</p>",
            unsafe_allow_html=True,
        )

    indice = opcoes_select.index(valor_atual) if valor_atual in opcoes_select else 0
    st.selectbox(
        label,
        options=opcoes_select,
        index=indice,
        key=chave_valor,
        label_visibility="collapsed",
    )
    valor_escolhido = st.session_state.get(chave_valor, "")
    st.session_state[chave_manual] = valor_escolhido != sugestao
    sugestao_texto = sugestao or "N/A"
    if cor:
        st.caption(
            f"<span style='color:{cor};'>Sugestão automática: {sugestao_texto}</span>",
            unsafe_allow_html=True,
        )
    else:
        st.caption(f"Sugestão automática: {sugestao_texto}")
    if st.session_state[chave_manual] and valor_escolhido:
        st.caption("Valor ajustado manualmente.")


def obter_bitola_cabo(
    distancia: float, potencia_kw: float, tabela=TABELA_BITOLAS
) -> str:
    """Retorna a bitola recomendada usando a tabela informada."""
    colunas = {
        1.9: "1,9kW",
        3.7: "3,7kW",
        7.4: "7,4kW",
        14.8: "2x7,4",
        11.0: "11,0kW",
        22.0: "22,0kW",
        44.0: "44,0kW",
    }
    if potencia_kw <= 0:
        return ""
    potencia_ref = min(colunas.keys(), key=lambda x: abs(x - potencia_kw))
    coluna = colunas[potencia_ref]
    linha = tabela[tabela["Distância (m)"] >= distancia]
    if linha.empty:
        return ""
    bitola = linha.iloc[0][coluna]
    return f"{bitola} mm²"


def dimensionar_disjuntor(instalacao: str, bitola_fase: str) -> str:
    """Sugere um disjuntor a partir da instalação e da bitola das fases."""
    match = re.search(r"(\d+(?:[.,]\d+)?)", str(bitola_fase))
    if not match:
        return ""
    bitola = float(match.group(1).replace(",", "."))
    mapa = {
        1.5: 16,
        2.5: 20,
        4.0: 25,
        6.0: 32,
        10.0: 40,
        16.0: 63,
        25.0: 80,
        35.0: 100,
        50.0: 125,
        70.0: 160,
    }
    corrente = mapa.get(bitola)
    if corrente is None:
        return ""
    polos = {
        "Monofásico": "1P+N",
        "Bifásico": "2P",
        "Trifásico": "3P",
    }.get(instalacao, "")
    return f"{polos} {corrente} A - DIN Curva C"


def dimensionar_idr(instalacao: str, bitola_fase: str) -> str:
    """Sugere um IDR de classe A a partir da instalação e da bitola das fases."""
    match = re.search(r"(\d+(?:[.,]\d+)?)", str(bitola_fase))
    if not match:
        return ""
    bitola = float(match.group(1).replace(",", "."))
    mapa = {
        1.5: 16,
        2.5: 20,
        4.0: 25,
        6.0: 32,
        10.0: 40,
        16.0: 63,
        25.0: 80,
        35.0: 100,
        50.0: 125,
        70.0: 160,
    }
    corrente = mapa.get(bitola)
    if corrente is None:
        return ""
    polos = {
        "Monofásico": "2P",
        "Bifásico": "2P",
        "Trifásico": "4P",
    }.get(instalacao, "")
    return f"{polos} {corrente} A - IDR Classe A 30 mA"


def dimensionar_dps(
    instalacao: str, bitola_fase: str, quantidade_carregadores: int
) -> str:
    """Sugere um DPS Tipo 2 de 1 polo a partir da instalação, bitola e quantidade."""
    match = re.search(r"(\d+(?:[.,]\d+)?)", str(bitola_fase))
    if not match:
        return ""
    bitola = float(match.group(1).replace(",", "."))
    mapa_corrente = {
        1.5: 16,
        2.5: 20,
        4.0: 25,
        6.0: 32,
        10.0: 40,
        16.0: 63,
        25.0: 80,
        35.0: 100,
        50.0: 125,
        70.0: 160,
    }
    corrente = mapa_corrente.get(bitola)
    if corrente is None:
        return ""
    if corrente <= 63:
        ka = 20
    elif corrente <= 125:
        ka = 40
    else:
        ka = 65
    condutores = {
        "Monofásico": 2,
        "Bifásico": 2,
        "Trifásico": 4,
    }.get(instalacao)
    if not condutores:
        return ""
    quantidade = condutores * max(1, quantidade_carregadores)
    return f"{quantidade}x 1P {ka} kA - DPS Tipo 2"


def render_dimensionamento_tab(tab_dimensionamento):
    """Renderiza a aba de Dimensionamento."""
    with tab_dimensionamento:
        exibir_tab_quadro_distribuicao = (
            st.session_state.get("quadro_distribuicao", "") == "Sim"
        )
        nomes_abas = ["CE Único"]
        if exibir_tab_quadro_distribuicao:
            nomes_abas.append("Quadro de Distribuição")
        nomes_abas.append("Tabelas Elétricas")
        abas_dimensionamento = st.tabs(nomes_abas)
        tab_resumo = abas_dimensionamento[0]
        tab_quadro_distribuicao = (
            abas_dimensionamento[1] if exibir_tab_quadro_distribuicao else None
        )
        tab_tabelas_eletricas = (
            abas_dimensionamento[2] if exibir_tab_quadro_distribuicao else abas_dimensionamento[1]
        )

        with tab_resumo:
            tipo_servico_dim = st.session_state.get("tipo_servico", "")
            tipo_local_dim = st.session_state.get("tipo_local", "")
            possui_carregador_dim = st.session_state.get("possui_carregador", "")
            quantidade_carregadores_dim = st.session_state.get("quantidade_carregadores", 0)
            potencia_carregador_dim = st.session_state.get("potencia_carregador", "")
            if potencia_carregador_dim == "Outro":
                val = st.session_state.get("pot_outro_valor", 0.0)
                potencia_kw = float(val)
                potencia_dim = f"{val} kW"
            else:
                match_pot = re.search(r"([\d.,]+)", str(potencia_carregador_dim))
                potencia_kw = (
                    float(match_pot.group(1).replace(",", "."))
                    if match_pot
                    else 0.0
                )
                potencia_dim = potencia_carregador_dim

            marca_carregadores_dim = st.session_state.get("marca_carregadores", "")
            tipo_conectividade_dim = st.session_state.get("tipo_conectividade", "")

            with st.expander("📋 Informações Gerais", expanded=True):
                cols = st.columns(7)
                with cols[0]:
                    st.markdown(f"**Serviço:** {tipo_servico_dim}")
                with cols[1]:
                    st.markdown(f"**Local:** {tipo_local_dim}")
                with cols[2]:
                    st.markdown(f"**Possui Carregador:** {possui_carregador_dim}")
                with cols[3]:
                    st.markdown(f"**Qtde:** {quantidade_carregadores_dim}")
                with cols[4]:
                    st.markdown(f"**Potência:** {potencia_dim}")
                with cols[5]:
                    st.markdown(f"**Marca:** {marca_carregadores_dim}")
                with cols[6]:
                    st.markdown(f"**OCPP:** {tipo_conectividade_dim}")

                percursos = st.session_state.get("percursos", [])
                distancia_total = sum(t for _, t in percursos)
                if percursos:
                    distancia_total_str = f"{distancia_total:g}"
                    st.markdown(
                        f"**Soma da Distância com Direções:** {distancia_total_str} m"
                    )

            observacoes_resumo = st.session_state.get("observacoes", "").strip()
            if observacoes_resumo:
                observacoes_formatadas = observacoes_resumo.replace("\n", "<br>")
                st.markdown("**Observações da Visita**", unsafe_allow_html=True)
                st.markdown(
                    f"<div style='margin-bottom: 1rem;'>{observacoes_formatadas}</div>",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown("**Observações da Visita**")
                st.markdown("_Nenhuma observação registrada._")

            with st.expander("📊 Dados do Quadro de Alimentação", expanded=False):
                alimentacao_dim = st.session_state.get("alimentacao", "")
                dj_opts = []
                if st.session_state.get("dj_disjuntor"):
                    dj_opts.append("Disjuntor")
                if st.session_state.get("dj_fusivel"):
                    dj_opts.append("Fusível")
                if st.session_state.get("dj_outro"):
                    dj_opts.append("Outro")
                dj_dim = ", ".join(dj_opts)

                corrente_disjuntor_dim = st.session_state.get("corrente_disjuntor", "")
                corrente_disjuntor_display = corrente_disjuntor_dim
                match_corrente = re.search(
                    r"(\d+\.?\d*)", str(corrente_disjuntor_dim).replace(",", ".")
                )
                if (
                    match_corrente and float(match_corrente.group(1)) < 63
                ) or (
                    isinstance(corrente_disjuntor_dim, str)
                    and "abaixo" in corrente_disjuntor_dim.lower()
                ):
                    corrente_disjuntor_display = (
                        f"<span style='color: red; font-weight: bold;'>{corrente_disjuntor_dim}</span>"
                    )
                bitola_cabos_dim = st.session_state.get("bitola_cabos", "")
                bitola_cabos_display = bitola_cabos_dim
                match_bitola = re.search(
                    r"(\d+\.?\d*)", str(bitola_cabos_dim).replace(",", ".")
                )
                if match_bitola and float(match_bitola.group(1)) < 10:
                    bitola_cabos_display = (
                        f"<span style='color: red; font-weight: bold;'>{bitola_cabos_dim}</span>"
                    )
                sistema_aterramento_dim = st.session_state.get("sistema_aterramento", "")
                sistema_aterramento_display = sistema_aterramento_dim
                if sistema_aterramento_dim in ["Sem aterramento", "TT", "IT"]:
                    sistema_aterramento_display = (
                        f"<span style='color: red; font-weight: bold;'>{sistema_aterramento_dim}</span>"
                    )
                barra_neutro_terra_dim = st.session_state.get("barra_neutro_terra", "")
                espaco_dj_saida_dim = st.session_state.get("espaco_dj_saida", "")

                tensoes_ff = {
                    "R/S": st.session_state.get("tensao_rs", ""),
                    "R/T": st.session_state.get("tensao_rt", ""),
                    "S/T": st.session_state.get("tensao_st", ""),
                }
                tensoes_fn = {
                    "R/N": st.session_state.get("tensao_rn", ""),
                    "S/N": st.session_state.get("tensao_sn", ""),
                    "T/N": st.session_state.get("tensao_tn", ""),
                }
                tensoes_ft = {
                    "R/T Terra": st.session_state.get("tensao_rtt", ""),
                    "S/T Terra": st.session_state.get("tensao_stt", ""),
                    "T/T Terra": st.session_state.get("tensao_ttt", ""),
                }
                tensao_nt_dim = st.session_state.get("tensao_n_t", "")
                correntes = {
                    "R": st.session_state.get("corrente_r", ""),
                    "S": st.session_state.get("corrente_s", ""),
                    "T": st.session_state.get("corrente_t", ""),
                }

                # Cálculo de corrente a partir da potência e média de tensões entre fases
                potencias_disponiveis = {
                    "0 kW": 0.0,
                    "1,9 kW": 1.9,
                    "3,7 kW": 3.7,
                    "7,4 kW": 7.4,
                    "11 kW": 11.0,
                    "22 kW": 22.0,
                    "44 kW": 44.0,
                }
                potencia_escolhida = st.session_state.get("potencia_carregador", "")
                if potencia_escolhida == "Outro":
                    potencia_kw = st.session_state.get("pot_outro_valor", 0.0)
                else:
                    potencia_kw = potencias_disponiveis.get(potencia_escolhida, 0.0)
                tensoes_ff_vals = []
                for val in tensoes_ff.values():
                    try:
                        tensoes_ff_vals.append(float(str(val).replace(',', '.')))
                    except (ValueError, TypeError):
                        pass
                media_tensao = (
                    sum(tensoes_ff_vals) / len(tensoes_ff_vals)
                    if tensoes_ff_vals
                    else 0.0
                )
                corrente_calc = ""
                if potencia_kw > 0 and media_tensao > 0:
                    if potencia_kw >= 11.0:
                        corrente_val = (potencia_kw * 1000) / (media_tensao * math.sqrt(3))
                    else:
                        corrente_val = (potencia_kw * 1000) / media_tensao
                    corrente_calc = f"{corrente_val:.2f}"
                st.session_state["corrente_calculada"] = corrente_calc

                bitola_calc = ""
                bitola_neutro_terra_calc = ""
                if distancia_total > 0 and potencia_kw > 0:
                    bitola_calc = obter_bitola_cabo(distancia_total, potencia_kw)
                    bitola_neutro_terra_calc = obter_bitola_cabo(
                        distancia_total, potencia_kw, TABELA_NEUTRO_TERRA
                    )
                _registrar_bitola_calculada(
                    "bitola_sugerida",
                    "bitola_sugerida_calculada",
                    "bitola_fase_manual",
                    bitola_calc,
                )
                _registrar_bitola_calculada(
                    "bitola_fase2_sugerida",
                    "bitola_fase2_sugerida_calculada",
                    "bitola_fase2_manual",
                    bitola_calc,
                )
                _registrar_bitola_calculada(
                    "bitola_fase3_sugerida",
                    "bitola_fase3_sugerida_calculada",
                    "bitola_fase3_manual",
                    bitola_calc,
                )
                _registrar_bitola_calculada(
                    "bitola_neutro_sugerida",
                    "bitola_neutro_sugerida_calculada",
                    "bitola_neutro_manual",
                    bitola_neutro_terra_calc,
                )
                _registrar_bitola_calculada(
                    "bitola_terra_sugerida",
                    "bitola_terra_sugerida_calculada",
                    "bitola_terra_manual",
                    bitola_neutro_terra_calc,
                )

                tensoes_fn_vals = []
                for val in tensoes_fn.values():
                    try:
                        tensoes_fn_vals.append(float(str(val).replace(',', '.')))
                    except (ValueError, TypeError):
                        pass
                is_monofasica_low_voltage = (
                    alimentacao_dim.strip().lower() == "monofásica"
                    and any(v < 210 for v in tensoes_fn_vals)
                )
                if is_monofasica_low_voltage:
                    alimentacao_display = (
                        "<span style='color: red; font-weight: bold;'>Monofásica</span>"
                    )
                else:
                    alimentacao_display = alimentacao_dim

                cols_visuais = st.columns(7)
                with cols_visuais[0]:
                    st.markdown(alimentacao_display, unsafe_allow_html=True)
                with cols_visuais[1]:
                    st.markdown(f"{dj_dim}")
                with cols_visuais[2]:
                    st.markdown(corrente_disjuntor_display, unsafe_allow_html=True)
                with cols_visuais[3]:
                    st.markdown(bitola_cabos_display, unsafe_allow_html=True)
                with cols_visuais[4]:
                    st.markdown(sistema_aterramento_display, unsafe_allow_html=True)
                with cols_visuais[5]:
                    st.markdown(f"{barra_neutro_terra_dim}")
                with cols_visuais[6]:
                    st.markdown(f"{espaco_dj_saida_dim}")

                st.markdown(", ".join([f"{k}: {v}" for k, v in tensoes_ff.items()]))
                st.markdown(", ".join([f"{k}: {v}" for k, v in tensoes_fn.items()]))
                st.markdown(
                    ", ".join([f"{k}: {v}" for k, v in tensoes_ft.items()])
                    + f", N/T: {tensao_nt_dim}"
                )
                st.markdown(", ".join([f"{k}: {v}" for k, v in correntes.items()]))

            st.markdown(
                """
                <style>
                input[disabled] {
                    color: #000 !important;
                    background-color: #ffffff !important;
                    opacity: 1 !important;
                    font-weight: bold !important;
                }
                div[data-testid=\"stSelectbox\"] div[role=\"combobox\"] {
                    font-size: 24px !important;
                    font-weight: bold !important;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                """
                <style>
                div[data-testid=\"stTextInput\"] input[aria-label=\"Corrente calculada (A)\"],
                div[data-testid=\"stTextInput\"] input[aria-label=\"Corrente Nominal (A)\"] {
                    font-size: 24px;
                    font-weight: bold;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )

            corrente_calculada_val = st.session_state.get("corrente_calculada", "")
            if "corrente_nominal_manual" not in st.session_state:
                st.session_state["corrente_nominal_manual"] = False
            corrente_nominal_manual = st.session_state.get("corrente_nominal_manual", False)
            corrente_nominal_atual = st.session_state.get("corrente_nominal")
            if corrente_nominal_atual is None:
                st.session_state["corrente_nominal"] = corrente_calculada_val
            elif (
                corrente_nominal_manual
                and st.session_state.get("corrente_nominal") == corrente_calculada_val
            ):
                st.session_state["corrente_nominal_manual"] = False
                corrente_nominal_manual = False
            if (
                not st.session_state.get("corrente_nominal_manual", False)
                and st.session_state.get("corrente_nominal") != corrente_calculada_val
            ):
                st.session_state["corrente_nominal"] = corrente_calculada_val

            cols_corrente = st.columns(2)
            with cols_corrente[0]:
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>Corrente calculada (A)</p>",
                    unsafe_allow_html=True,
                )
                st.text_input(
                    "Corrente calculada (A)",
                    value=corrente_calculada_val,
                    key="corrente_calculada",
                    disabled=True,
                    label_visibility="collapsed",
                )
            with cols_corrente[1]:
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>Corrente Nominal (A)</p>",
                    unsafe_allow_html=True,
                )
                st.text_input(
                    "Corrente Nominal (A)",
                    key="corrente_nominal",
                    label_visibility="collapsed",
                    on_change=_atualizar_corrente_nominal_manual,
                )
            tensoes = []
            for campo in ("tensao_rs", "tensao_rt", "tensao_st"):
                try:
                    tensoes.append(
                        float(str(st.session_state.get(campo, "")).replace(",", "."))
                    )
                except (ValueError, TypeError):
                    pass
            instalacao_sugerida = st.session_state.get("instalacao_sistema", "Monofásico")
            if potencia_kw >= 11:
                instalacao_sugerida = "Trifásico"
            elif potencia_kw <= 7.4 and len(tensoes) == 3:
                if all(t > 360 for t in tensoes):
                    instalacao_sugerida = "Monofásico"
                elif all(210 <= t <= 240 for t in tensoes):
                    instalacao_sugerida = "Bifásico"
            st.session_state["instalacao_sistema"] = instalacao_sugerida
            with st.expander("🧵 Dimensionamento dos Cabos", expanded=False):
                col_instalacao, col_tipo = st.columns(2)
                with col_instalacao:
                    instalacao_sistema = st.radio(
                        "Instalação em Sistema:",
                        ["Monofásico", "Bifásico", "Trifásico"],
                        key="instalacao_sistema",
                    )
                with col_tipo:
                    tipo_cabos = st.radio(
                        "Tipo de Cabos",
                        ["Cabo PVC", "Cabo HEPR"],
                        key="tipo_cabos",
                    )
                tabela_key, tabela_base, _ = _obter_configuracao_tabela_cabos(
                    tipo_cabos
                )
                tabela_cabos = _normalizar_tabela_cabos(
                    st.session_state.get(tabela_key, tabela_base.copy()),
                    tabela_base,
                )
                # Reinicia a tabela do outro tipo de cabo quando houver mudança de seleção
                previous_tipo = st.session_state.get("tipo_cabos_prev")
                if previous_tipo and previous_tipo != tipo_cabos:
                    if tipo_cabos == "Cabo PVC":
                        st.session_state["tabela_cabos_hepr"] = (
                            TABELA_CABO_UNIPOLAR_HEPR.copy()
                        )
                        st.session_state.pop("tabela_cabos_editor_hepr", None)
                    else:
                        st.session_state["tabela_cabos_pvc"] = (
                            TABELA_CABO_ISOLADO_PVC.copy()
                        )
                        st.session_state.pop("tabela_cabos_editor_pvc", None)
                    st.session_state["area_total_ocupada"] = 0.0
                st.session_state["tipo_cabos_prev"] = tipo_cabos
                bitola_label = (
                    "Bitola Cabo Preto PVC fase 1 (mm²)"
                    if tipo_cabos == "Cabo PVC"
                    else "Bitola Cabo Preto HEPR fase 1 (mm²)"
                )
                bitola_label_fase2 = (
                    "Bitola Cabo Preto PVC fase 2 (mm²)"
                    if tipo_cabos == "Cabo PVC"
                    else "Bitola Cabo Preto HEPR fase 2 (mm²)"
                )
                bitola_label_fase3 = (
                    "Bitola Cabo Preto PVC fase 3 (mm²)"
                    if tipo_cabos == "Cabo PVC"
                    else "Bitola Cabo Preto HEPR fase 3 (mm²)"
                )
                bitola_label_neutro = (
                    "Bitola Cabo Azul PVC para Neutro (mm²)"
                    if tipo_cabos == "Cabo PVC"
                    else "Bitola Cabo Azul HEPR para Neutro (mm²)"
                )
                bitola_label_terra = (
                    "Bitola Cabo Verde PVC para Terra (mm²)"
                    if tipo_cabos == "Cabo PVC"
                    else "Bitola Cabo Verde HEPR para Terra (mm²)"
                )
                opcoes_bitola = _obter_opcoes_bitola_cabos(tabela_cabos)
                _render_bitola_select(
                    bitola_label,
                    "bitola_sugerida",
                    "bitola_sugerida_calculada",
                    "bitola_fase_manual",
                    opcoes_bitola,
                )
                if instalacao_sistema == "Monofásico":
                    _render_bitola_select(
                        bitola_label_neutro,
                        "bitola_neutro_sugerida",
                        "bitola_neutro_sugerida_calculada",
                        "bitola_neutro_manual",
                        opcoes_bitola,
                        cor="#00A1DA",
                    )
                elif instalacao_sistema == "Bifásico":
                    st.session_state["bitola_neutro_sugerida"] = ""
                    st.session_state["bitola_neutro_sugerida_calculada"] = ""
                    st.session_state["bitola_neutro_manual"] = False
                    _render_bitola_select(
                        bitola_label_fase2,
                        "bitola_fase2_sugerida",
                        "bitola_fase2_sugerida_calculada",
                        "bitola_fase2_manual",
                        opcoes_bitola,
                    )
                else:  # Trifásico
                    _render_bitola_select(
                        bitola_label_fase2,
                        "bitola_fase2_sugerida",
                        "bitola_fase2_sugerida_calculada",
                        "bitola_fase2_manual",
                        opcoes_bitola,
                    )
                    _render_bitola_select(
                        bitola_label_fase3,
                        "bitola_fase3_sugerida",
                        "bitola_fase3_sugerida_calculada",
                        "bitola_fase3_manual",
                        opcoes_bitola,
                    )
                    _render_bitola_select(
                        bitola_label_neutro,
                        "bitola_neutro_sugerida",
                        "bitola_neutro_sugerida_calculada",
                        "bitola_neutro_manual",
                        opcoes_bitola,
                        cor="#00A1DA",
                    )
                _render_bitola_select(
                    bitola_label_terra,
                    "bitola_terra_sugerida",
                    "bitola_terra_sugerida_calculada",
                    "bitola_terra_manual",
                    opcoes_bitola,
                    cor="#008000",
                )

        with tab_tabelas_eletricas:
            st.subheader("📘 Tabela de Fases")
            st.dataframe(TABELA_BITOLAS)
            st.subheader("📗 Tabela de Seção dos fios Neutro e Terra")
            st.dataframe(TABELA_NEUTRO_TERRA)

            st.subheader("📕 Tabela de Eletrodutos")
            st.dataframe(TABELA_ELETRODUTOS)

            tipo_cabos = st.session_state.get("tipo_cabos", "Cabo PVC")
            tabela_key, tabela_base, editor_key = _obter_configuracao_tabela_cabos(
                tipo_cabos
            )
            if tipo_cabos == "Cabo PVC":
                st.subheader(
                    "📙 Nº1 (Cabo Isolado PVC) Flexicom Antichama 450/750 V 70°C - Classe 4 ou 5"
                )
            else:
                st.subheader(
                    "📒 Nº6 (Cabo Unipolar HEPR) GTEPROM Flex 90°C Antichama 0,6/1kV - Classe 5"
                )
            tabela_cabos = _normalizar_tabela_cabos(
                st.session_state.get(tabela_key, tabela_base.copy()), tabela_base
            )
            # Garante que a coluna de quantidade exista e mantenha valores já editados
            if "Quantidade de Cabos" not in tabela_cabos.columns:
                tabela_cabos["Quantidade de Cabos"] = 0
            else:
                tabela_cabos["Quantidade de Cabos"] = pd.to_numeric(
                    tabela_cabos["Quantidade de Cabos"], errors="coerce"
                ).fillna(0)

            # Preenche automaticamente a quantidade de cabos com base nas
            # bitolas recomendadas para fases, neutro e terra.
            def _parse_bitola(valor):
                match = re.search(r"(\d+(?:[.,]\d+)?)", str(valor))
                if not match:
                    return None
                return float(match.group(1).replace(",", "."))

            bitola_fase = _parse_bitola(st.session_state.get("bitola_sugerida"))
            bitola_neutro = _parse_bitola(
                st.session_state.get("bitola_neutro_sugerida")
            )
            bitola_terra = _parse_bitola(
                st.session_state.get("bitola_terra_sugerida")
            )
            quantidade_fases = {"Monofásico": 1, "Bifásico": 2, "Trifásico": 3}.get(
                instalacao_sistema, 0
            )
            quantidades = {}
            if bitola_fase is not None:
                quantidades[bitola_fase] = quantidades.get(bitola_fase, 0) + quantidade_fases
            if bitola_neutro is not None and instalacao_sistema != "Bifásico":
                quantidades[bitola_neutro] = quantidades.get(bitola_neutro, 0) + 1
            if bitola_terra is not None:
                quantidades[bitola_terra] = quantidades.get(bitola_terra, 0) + 1

            for bitola, qtd in quantidades.items():
                mask = tabela_cabos["Cabo (mm²)"] == bitola
                tabela_cabos.loc[mask, "Quantidade de Cabos"] = tabela_cabos.loc[mask,
                    "Quantidade de Cabos"].replace(0, qtd)

            tabela_cabos["Área do Cabo (mm²)"] = (
                math.pi * (tabela_cabos["Diâmetro Externo (mm)"] / 2) ** 2
            )
            tabela_cabos["Área Ocupável Condutores"] = (
                tabela_cabos["Área do Cabo (mm²)"]
                * tabela_cabos["Quantidade de Cabos"]
            )
            tabela_cabos = st.data_editor(
                tabela_cabos,
                key=editor_key,
                column_config={
                    "Quantidade de Cabos": st.column_config.NumberColumn(
                        "Quantidade de Cabos", min_value=0
                    ),
                    "Área do Cabo (mm²)": st.column_config.NumberColumn(
                        "Área do Cabo (mm²)", format="%.8f", disabled=True
                    ),
                    "Área Ocupável Condutores": st.column_config.NumberColumn(
                        "Área Ocupável Condutores", format="%.7f", disabled=True
                    ),
                },
            )
            tabela_cabos["Área Ocupável Condutores"] = (
                tabela_cabos["Área do Cabo (mm²)"]
                * tabela_cabos["Quantidade de Cabos"]
            )
            st.session_state[tabela_key] = tabela_cabos
            area_total_ocupada = tabela_cabos["Área Ocupável Condutores"].sum()
            st.session_state["area_total_ocupada"] = area_total_ocupada
            st.markdown(
                """
                <style>
                div[data-testid="stNumberInput"] input[aria-label="Área Total Ocupada"] {
                    font-size: 20px;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                "<p style='font-size:20px;'>Área Total Ocupada</p>",
                unsafe_allow_html=True,
            )
            st.number_input(
                "Área Total Ocupada",
                value=float(area_total_ocupada),
                format="%.7f",
                disabled=True,
                label_visibility="collapsed",
            )
            tamanho_eletroduto = ""
            if area_total_ocupada > 0:
                linha_eletroduto = TABELA_ELETRODUTOS[
                    TABELA_ELETRODUTOS["Área Ocupável 40% (mm²)"]
                    >= area_total_ocupada
                ]
                if not linha_eletroduto.empty:
                    tamanho_eletroduto = linha_eletroduto.iloc[0][
                        "Eletroduto (Pol)"
                    ]

            st.session_state["tamanho_eletroduto_sugerido"] = tamanho_eletroduto
            tamanho_atual = st.session_state.get("tamanho_eletroduto", "")

            if tamanho_atual == tamanho_eletroduto:
                st.session_state["tamanho_eletroduto_manual"] = False

            if not st.session_state.get("tamanho_eletroduto_manual", False) or not tamanho_atual:
                st.session_state["tamanho_eletroduto"] = tamanho_eletroduto
                st.session_state["tamanho_eletroduto_resumo"] = tamanho_eletroduto
            else:
                st.session_state.setdefault(
                    "tamanho_eletroduto_resumo", tamanho_atual
                )
            st.markdown(
                 """
                 <style>
                 div[data-testid="stTextInput"] input[aria-label="Tamanho do Eletroduto"] {
                     font-size: 24px;
                     font-weight: bold;
                    background-color: #ffffff;
                    color: #000000;
                 }
                 </style>
                 """,
                 unsafe_allow_html=True,
             )
            st.markdown(
                "<p style='font-size:24px; font-weight:bold;'>Tamanho do Eletroduto</p>",
                unsafe_allow_html=True,
            )
            st.text_input(
                "Tamanho do Eletroduto",
                value=st.session_state.get("tamanho_eletroduto", tamanho_eletroduto),
                key="tamanho_eletroduto",
                disabled=True,
                label_visibility="collapsed",
            )

        if tab_quadro_distribuicao is not None:
            with tab_quadro_distribuicao:
                percursos_quadro = st.session_state.get("percursos_quadro", [])
                total_quadro = sum(trecho for _, trecho in percursos_quadro)
                st.session_state["distancia_alimentacao_distribuicao_tecnica"] = (
                    f"{total_quadro:g}"
                )

                with st.expander("🛠️ Informações Técnicas", expanded=False):
                    st.text_input(
                        "Distãncia entre Alimentação e Distribuição",
                        value=st.session_state[
                            "distancia_alimentacao_distribuicao_tecnica"
                        ],
                        key="distancia_alimentacao_distribuicao_tecnica",
                        disabled=True,
                    )

                    st.markdown("**Medições da aba Visita**")

                    col_rs, col_rt, col_st = st.columns(3)
                    with col_rs:
                        st.text_input(
                            "Tensão R/S",
                            value=st.session_state.get("tensao_rs", ""),
                            disabled=True,
                        )
                    with col_rt:
                        st.text_input(
                            "Tensão R/T",
                            value=st.session_state.get("tensao_rt", ""),
                            disabled=True,
                        )
                    with col_st:
                        st.text_input(
                            "Tensão S/T",
                            value=st.session_state.get("tensao_st", ""),
                            disabled=True,
                        )

                    col_rn, col_sn, col_tn = st.columns(3)
                    with col_rn:
                        st.text_input(
                            "Tensão R/N",
                            value=st.session_state.get("tensao_rn", ""),
                            disabled=True,
                        )
                    with col_sn:
                        st.text_input(
                            "Tensão S/N",
                            value=st.session_state.get("tensao_sn", ""),
                            disabled=True,
                        )
                    with col_tn:
                        st.text_input(
                            "Tensão T/N",
                            value=st.session_state.get("tensao_tn", ""),
                            disabled=True,
                        )

                    col_rtt, col_stt, col_ttt = st.columns(3)
                    with col_rtt:
                        st.text_input(
                            "Tensão R/T Terra",
                            value=st.session_state.get("tensao_rtt", ""),
                            disabled=True,
                        )
                    with col_stt:
                        st.text_input(
                            "Tensão S/T Terra",
                            value=st.session_state.get("tensao_stt", ""),
                            disabled=True,
                        )
                    with col_ttt:
                        st.text_input(
                            "Tensão T/T Terra",
                            value=st.session_state.get("tensao_ttt", ""),
                            disabled=True,
                        )

                    st.text_input(
                        "Tensão N/T",
                        value=st.session_state.get("tensao_n_t", ""),
                        disabled=True,
                    )

                    col_r, col_s, col_t = st.columns(3)
                    with col_r:
                        st.text_input(
                            "Corrente R",
                            value=st.session_state.get("corrente_r", ""),
                            disabled=True,
                        )
                    with col_s:
                        st.text_input(
                            "Corrente S",
                            value=st.session_state.get("corrente_s", ""),
                            disabled=True,
                        )
                    with col_t:
                        st.text_input(
                            "Corrente T",
                            value=st.session_state.get("corrente_t", ""),
                            disabled=True,
                        )

                with st.expander("🧰 Painel", expanded=False):
                    st.markdown("_Sem informações de painel cadastradas._")

        with tab_resumo:
            with st.expander("\U0001F4D0 Dimensionamento da Infra-Seca", expanded=False):
                st.session_state[
                    "dimensionamento_infra_seca_resumo"
                ] = st.session_state.get("dimensionamento_infra_seca", "")
                st.markdown(
                    """
                    <style>
                    div[data-testid=\"stTextInput\"] input[aria-label=\"Tamanho do Eletroduto\"] {
                        font-size: 24px;
                        font-weight: bold;
                        background-color: #ffffff;
                        color: #000000;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>Tamanho do Eletroduto</p>",
                    unsafe_allow_html=True,
                )
                if "tamanho_eletroduto_manual" not in st.session_state:
                    st.session_state["tamanho_eletroduto_manual"] = False
                st.session_state.setdefault(
                    "tamanho_eletroduto_resumo",
                    st.session_state.get("tamanho_eletroduto", ""),
                )
                opcoes_eletroduto = [
                    ""
                ] + [
                    str(valor)
                    for valor in TABELA_ELETRODUTOS["Eletroduto (Pol)"].dropna().unique()
                ]
                valor_atual_eletroduto = str(
                    st.session_state.get("tamanho_eletroduto_resumo", "")
                )
                if valor_atual_eletroduto not in opcoes_eletroduto:
                    opcoes_eletroduto.append(valor_atual_eletroduto)
                indice_eletroduto = opcoes_eletroduto.index(valor_atual_eletroduto)
                st.selectbox(
                    "Tamanho do Eletroduto",
                    options=opcoes_eletroduto,
                    index=indice_eletroduto,
                    key="tamanho_eletroduto_resumo",
                    on_change=_sincronizar_tamanho_eletroduto_personalizado,
                    label_visibility="collapsed",
                )
                tamanho_sugerido = st.session_state.get("tamanho_eletroduto_sugerido", "")
                if tamanho_sugerido:
                    st.caption(f"Sugestão automática: {tamanho_sugerido}")

            with st.expander("🛡️ Quadro de Proteção", expanded=False):
                disjuntor_val = dimensionar_disjuntor(
                    st.session_state.get("instalacao_sistema", ""),
                    st.session_state.get("bitola_sugerida", ""),
                )
                st.session_state["disjuntor_recomendado"] = disjuntor_val
                if "disjuntor_manual" not in st.session_state:
                    st.session_state["disjuntor_manual"] = False
                disjuntor_resumo_atual = st.session_state.get("disjuntor_resumo", "")
                if disjuntor_resumo_atual:
                    st.session_state["disjuntor_manual"] = disjuntor_resumo_atual != (
                        disjuntor_val or ""
                    )
                else:
                    st.session_state["disjuntor_manual"] = False
                if not st.session_state["disjuntor_manual"]:
                    st.session_state["disjuntor_resumo"] = disjuntor_val or ""
                idr_val = dimensionar_idr(
                    st.session_state.get("instalacao_sistema", ""),
                    st.session_state.get("bitola_sugerida", ""),
                )
                st.session_state["idr_recomendado"] = idr_val
                if "idr_manual" not in st.session_state:
                    st.session_state["idr_manual"] = False
                if not st.session_state["idr_manual"]:
                    st.session_state["idr_resumo"] = idr_val or ""
                dps_val = dimensionar_dps(
                    st.session_state.get("instalacao_sistema", ""),
                    st.session_state.get("bitola_sugerida", ""),
                    st.session_state.get("quantidade_carregadores", 1),
                )
                st.session_state["dps_recomendado"] = dps_val
                if "dps_manual" not in st.session_state:
                    st.session_state["dps_manual"] = False
                if not st.session_state["dps_manual"]:
                    st.session_state["dps_resumo"] = dps_val or ""
                barra_pente_map = {
                    "Monofásico": "Barra Pente Monopolar",
                    "Bifásico": "Barra Pente Bipolar",
                    "Trifásico": "Barra Pente Tripolar",
                }
                barra_pente_val = barra_pente_map.get(
                    st.session_state.get("instalacao_sistema", ""), ""
                )
                st.session_state["barra_pente_recomendado"] = barra_pente_val
                if "barra_pente_manual" not in st.session_state:
                    st.session_state["barra_pente_manual"] = False
                if not st.session_state["barra_pente_manual"]:
                    st.session_state["barra_pente_resumo"] = barra_pente_val or ""
                st.markdown(
                    """
                    <style>
                    div[data-testid=\"stSelectbox\"] div[role=\"combobox\"] {
                        font-size: 24px;
                        font-weight: bold;
                    }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>Disjuntor</p>",
                    unsafe_allow_html=True,
                )
                opcoes_disjuntor = _obter_opcoes_material(
                    "valores_disjuntores_din_df", "valores_disjuntor_din.csv"
                )
                _render_selectbox_material(
                    "Disjuntor",
                    "disjuntor_resumo",
                    opcoes_disjuntor,
                    disjuntor_val or "",
                )
                st.session_state["disjuntor_manual"] = (
                    st.session_state.get("disjuntor_resumo", "")
                    != (disjuntor_val or "")
                )
                if disjuntor_val:
                    st.caption(f"Sugestão automática: {disjuntor_val}")
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>IDR (Interruptor Diferencial Residual)</p>",
                    unsafe_allow_html=True,
                )
                opcoes_idr = _obter_opcoes_material("valores_idr_df", "valores_idr.csv")
                _render_selectbox_material(
                    "IDR (Interruptor Diferencial Residual)",
                    "idr_resumo",
                    opcoes_idr,
                    idr_val or "",
                )
                st.session_state["idr_manual"] = (
                    st.session_state.get("idr_resumo", "") != (idr_val or "")
                )
                if idr_val:
                    st.caption(f"Sugestão automática: {idr_val}")
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>DPS (Dispositivo de Proteção contra Surtos)</p>",
                    unsafe_allow_html=True,
                )
                opcoes_dps = _obter_opcoes_material("valores_dps_df", "valores_dps.csv")
                _render_selectbox_material(
                    "DPS (Dispositivo de Proteção contra Surtos)",
                    "dps_resumo",
                    opcoes_dps,
                    dps_val or "",
                )
                st.session_state["dps_manual"] = (
                    st.session_state.get("dps_resumo", "") != (dps_val or "")
                )
                if dps_val:
                    st.caption(f"Sugestão automática: {dps_val}")
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>Barra Pente</p>",
                    unsafe_allow_html=True,
                )
                opcoes_barra = _obter_opcoes_material(
                    "valores_barra_pente_df", "valores_barra_pente.csv"
                )
                _render_selectbox_material(
                    "Barra Pente",
                    "barra_pente_resumo",
                    opcoes_barra,
                    barra_pente_val or "",
                )
                st.session_state["barra_pente_manual"] = (
                    st.session_state.get("barra_pente_resumo", "")
                    != (barra_pente_val or "")
                )
                if barra_pente_val:
                    st.caption(f"Sugestão automática: {barra_pente_val}")
                st.markdown(
                    "<p style='font-size:24px; font-weight:bold;'>Tipo de Quadro</p>",
                    unsafe_allow_html=True,
                )
                tipo_quadro = st.selectbox(
                    "Tipo de Quadro",
                    ["PVC", "Metálico"],
                    key="tipo_quadro_resumo",
                    label_visibility="collapsed",
                )
                if tipo_quadro == "PVC":
                    quadro_pvc_map = {
                        "Monofásico": "Quadro PVC de 8 Posições",
                        "Bifásico": "Quadro PVC de 8 Posições",
                        "Trifásico": "Quadro PVC de 12 Posições",
                    }
                    quadro_pvc_val = quadro_pvc_map.get(
                        st.session_state.get("instalacao_sistema", ""), ""
                    )
                    st.session_state["quadro_pvc_recomendado"] = quadro_pvc_val
                    st.session_state["quadro_metalico_recomendado"] = ""
                    if "quadro_pvc_manual" not in st.session_state:
                        st.session_state["quadro_pvc_manual"] = False
                    valor_quadro_pvc_atual = st.session_state.get(
                        "quadro_pvc_resumo", ""
                    )
                    if (
                        not st.session_state["quadro_pvc_manual"]
                        and valor_quadro_pvc_atual
                        and valor_quadro_pvc_atual != (quadro_pvc_val or "")
                    ):
                        st.session_state["quadro_pvc_manual"] = True
                    if not st.session_state["quadro_pvc_manual"]:
                        st.session_state["quadro_pvc_resumo"] = quadro_pvc_val or ""
                    st.markdown(
                        "<p style='font-size:24px; font-weight:bold;'>Quadro PVC</p>",
                        unsafe_allow_html=True,
                    )
                    opcoes_quadros = _obter_opcoes_material(
                        "valores_paineis_quadros_df", "valores_paineis_quadros.csv"
                    )
                    opcoes_pvc = [
                        opt
                        for opt in opcoes_quadros
                        if not opt or "pvc" in opt.lower()
                    ]
                    if not opcoes_pvc:
                        opcoes_pvc = opcoes_quadros
                    _render_selectbox_material(
                        "Quadro PVC",
                        "quadro_pvc_resumo",
                        opcoes_pvc,
                        quadro_pvc_val or "",
                    )
                    st.session_state["quadro_pvc_manual"] = (
                        st.session_state.get("quadro_pvc_resumo", "")
                        != (quadro_pvc_val or "")
                    )
                    if quadro_pvc_val:
                        st.caption(f"Sugestão automática: {quadro_pvc_val}")
                elif tipo_quadro == "Metálico":
                    quadro_metalico_map = {
                        "Monofásico": "Quadro 20x20x14",
                        "Bifásico": "Quadro 20x20x14",
                        "Trifásico": "Quadro 30x30x20",
                    }
                    quadro_metalico_val = quadro_metalico_map.get(
                        st.session_state.get("instalacao_sistema", ""), ""
                    )
                    st.session_state["quadro_metalico_recomendado"] = quadro_metalico_val
                    st.session_state["quadro_pvc_recomendado"] = ""
                    if "quadro_metalico_manual" not in st.session_state:
                        st.session_state["quadro_metalico_manual"] = False
                    valor_quadro_metalico_atual = st.session_state.get(
                        "quadro_metalico_resumo", ""
                    )
                    if (
                        not st.session_state["quadro_metalico_manual"]
                        and valor_quadro_metalico_atual
                        and valor_quadro_metalico_atual != (quadro_metalico_val or "")
                    ):
                        st.session_state["quadro_metalico_manual"] = True
                    if not st.session_state["quadro_metalico_manual"]:
                        st.session_state["quadro_metalico_resumo"] = (
                            quadro_metalico_val or ""
                        )
                    st.markdown(
                        "<p style='font-size:24px; font-weight:bold;'>Quadro Metálico</p>",
                        unsafe_allow_html=True,
                    )
                    opcoes_quadros = _obter_opcoes_material(
                        "valores_paineis_quadros_df", "valores_paineis_quadros.csv"
                    )
                    opcoes_metal = [
                        opt
                        for opt in opcoes_quadros
                        if not opt or "metal" in opt.lower()
                    ]
                    if not opcoes_metal:
                        opcoes_metal = opcoes_quadros
                    _render_selectbox_material(
                        "Quadro Metálico",
                        "quadro_metalico_resumo",
                        opcoes_metal,
                        quadro_metalico_val or "",
                    )
                    st.session_state["quadro_metalico_manual"] = (
                        st.session_state.get("quadro_metalico_resumo", "")
                        != (quadro_metalico_val or "")
                    )
                    if quadro_metalico_val:
                        st.caption(f"Sugestão automática: {quadro_metalico_val}")
                else:
                    st.session_state["quadro_pvc_recomendado"] = ""
                    st.session_state["quadro_metalico_recomendado"] = ""

            with st.expander("\U0001F4E6 Material Adicional", expanded=False):
                if st.session_state.get("disjuntor_caixa_moldada") == "Sim":
                    df_dj = pd.read_csv(
                        "valores_disjuntor_caixa_moldada.csv", sep=";"
                    )
                    modelos_dj = [""] + df_dj["Modelo"].dropna().tolist()
                    modelo_key = "dimensionamento_modelo_disjuntor_caixa_moldada"
                    if modelo_key not in st.session_state:
                        st.session_state[modelo_key] = st.session_state.get(
                            "modelo_disjuntor_caixa_moldada", ""
                        )
                    st.selectbox(
                        "Disjuntor Caixa Moldada",
                        modelos_dj,
                        key=modelo_key,
                    )
                    st.session_state["modelo_disjuntor_caixa_moldada"] = st.session_state[
                        modelo_key
                    ]

                if st.session_state.get("tem_tomada_industrial") == "Sim":
                    opcoes_tomadas = [
                        "",
                        "Tomada Sobrepor Industrial 16A 2P+T 6h (Azul)",
                        "Tomada Sobrepor Industrial 32A 2P+T 6h (Azul)",
                        "Tomada Sobrepor 3p+T+N 32A 6h (Verm.)",
                        "Não sabe qual",
                    ]
                    st.selectbox(
                        "Tomada industrial",
                        opcoes_tomadas,
                        key="tomada_industrial",
                    )
                elif st.session_state.get("tomada_industrial"):
                    st.session_state["tomada_industrial"] = ""

                if st.session_state.get("tem_medidor") == "Sim":
                    opcoes_medidores = [
                        "",
                        "Medidor Bipolar Wifi",
                        "Medidor Bipolar",
                        "Outro",
                    ]
                    st.selectbox(
                        "Medidor",
                        opcoes_medidores,
                        key="medidor",
                    )
                elif st.session_state.get("medidor"):
                    st.session_state["medidor"] = ""

                if st.session_state.get("transformador") == "Sim":
                    produtos_transformadores = st.session_state.get(
                        "transformadores_produtos", []
                    )
                    if not produtos_transformadores:
                        produtos_transformadores = (
                            obter_produtos_transformadores_padrao()
                        )
                    opcoes_transformadores = [""] + produtos_transformadores
                    if (
                        st.session_state.get("transformador_produto", "")
                        not in opcoes_transformadores
                    ):
                        st.session_state["transformador_produto"] = ""
                    st.selectbox(
                        "Transformador",
                        opcoes_transformadores,
                        key="transformador_produto",
                    )
                else:
                    st.session_state["transformador_produto"] = ""

                if st.session_state.get("barra_roscada") == "Sim":
                    try:
                        df_barra_roscada = pd.read_csv(
                            "valores_barra_roscada.csv", sep=";"
                        )
                    except FileNotFoundError:
                        df_barra_roscada = pd.DataFrame(
                            [["Barra Roscada 1/4", "R$ 16,00", ""]],
                            columns=["Material", "Preco", "Atualizado"],
                        )

                    opcoes_barra = []
                    for _, row in df_barra_roscada.iterrows():
                        material = str(row.get("Material", "")).strip()
                        preco = str(row.get("Preco", "")).strip()
                        if not material:
                            continue
                        display = material if not preco else f"{material} - {preco}"
                        opcoes_barra.append(display)

                    if not opcoes_barra:
                        st.info("Nenhuma opção de barra roscada cadastrada no momento.")
                    else:
                        opcoes_barra = [""] + opcoes_barra
                        if (
                            st.session_state.get("barra_roscada_material", "")
                            not in opcoes_barra
                        ):
                            st.session_state["barra_roscada_material"] = ""
                        st.selectbox(
                            "Barra Roscada (vergalhão)",
                            opcoes_barra,
                            key="barra_roscada_material",
                        )
                else:
                    st.session_state["barra_roscada_material"] = ""

                if st.session_state.get("eletrocalha") == "Sim":
                    metros_eletrocalha = st.session_state.get("metros_eletrocalha", 0)
                    try:
                        metros_formatado = f"{float(metros_eletrocalha):g}"
                    except (TypeError, ValueError):
                        metros_formatado = str(metros_eletrocalha)
                    label_dimensoes = (
                        f"Dimensões Eletrocalha ({metros_formatado} m selecionados)"
                    )
                    opcoes_dimensoes = [
                        "Eletrocalha perfurada #24 50x50mm",
                        "Eletrocalha perfurada #24 75x50mm",
                        "Eletrocalha perfurada #24 75x75mm",
                        "Eletrocalha perfurada #24 100x75mm",
                        "Eletrocalha perfurada #24 100x100mm",
                        "Eletrocalha perfurada #24 150x100mm",
                        "Eletrocalha perfurada #24 200x100mm",
                        "Eletrocalha perfurada #24 250x100mm",
                        "Eletrocalha perfurada #24 300x100mm",
                    ]
                    dimensao_atual = st.session_state.get(
                        "dimensoes_eletrocalha", ""
                    )
                    opcoes_dimensoes_widget = [""] + opcoes_dimensoes
                    if (
                        dimensao_atual
                        and dimensao_atual not in opcoes_dimensoes_widget
                    ):
                        opcoes_dimensoes_widget.append(dimensao_atual)

                    try:
                        index_dimensao = opcoes_dimensoes_widget.index(dimensao_atual)
                    except ValueError:
                        index_dimensao = 0

                    dimensao_selecionada = st.selectbox(
                        label_dimensoes,
                        opcoes_dimensoes_widget,
                        index=index_dimensao,
                        key="dimensionamento_dimensoes_eletrocalha",
                    )
                    st.session_state["dimensoes_eletrocalha"] = dimensao_selecionada
                else:
                    st.session_state["dimensoes_eletrocalha"] = ""

            if st.button("Salvar Dados de Dimensionamento"):
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    load_workbook = None

                docs_dir = Path(__file__).with_name("Docs Salvos")
                docs_dir.mkdir(exist_ok=True)

                ordem_venda = st.session_state.get("ordem_venda", "")
                base_name = "Dados de Dimensionamento"
                filename = (
                    f"{base_name} {ordem_venda}.xlsx"
                    if ordem_venda
                    else f"{base_name}.xlsx"
                )
                filepath = docs_dir / filename

                dados_dimensionamento = {
                    "Ordem de Venda": ordem_venda,
                    "Cliente": st.session_state.get("cliente", ""),
                    "Tipo de Serviço": tipo_servico_dim,
                    "Tipo de Local": tipo_local_dim,
                    "Possui Carregador": possui_carregador_dim,
                    "Quantidade de Carregadores": quantidade_carregadores_dim,
                    "Potência Selecionada": potencia_dim,
                    "Potência (kW)": potencia_kw,
                    "Marca dos Carregadores": marca_carregadores_dim,
                    "Conectividade": tipo_conectividade_dim,
                    "Soma Distância (m)": distancia_total,
                    "Dimensionamento Infra-Seca": st.session_state.get(
                        "dimensionamento_infra_seca",
                        st.session_state.get("dimensionamento_infra_seca_resumo", ""),
                    ),
                    "Alimentação Declarada": alimentacao_dim,
                    "Instalação Sugerida": st.session_state.get("instalacao_sistema", ""),
                    "Dispositivo de Proteção": dj_dim,
                    "Corrente Disjuntor Informada": corrente_disjuntor_dim,
                    "Corrente Calculada (A)": st.session_state.get(
                        "corrente_calculada", ""
                    ),
                    "Bitola Cabos Informada": bitola_cabos_dim,
                    "Bitola Sugerida Fase 1": st.session_state.get(
                        "bitola_sugerida", ""
                    ),
                    "Bitola Sugerida Fase 2": st.session_state.get(
                        "bitola_fase2_sugerida", ""
                    ),
                    "Bitola Sugerida Fase 3": st.session_state.get(
                        "bitola_fase3_sugerida", ""
                    ),
                    "Bitola Sugerida Neutro": st.session_state.get(
                        "bitola_neutro_sugerida", ""
                    ),
                    "Bitola Sugerida Terra": st.session_state.get(
                        "bitola_terra_sugerida", ""
                    ),
                    "Sistema de Aterramento": sistema_aterramento_dim,
                    "Barra Neutro/Terra": barra_neutro_terra_dim,
                    "Espaço DJ Saída": espaco_dj_saida_dim,
                    "Tensões Fase-Fase": ", ".join(
                        f"{k}: {v}" for k, v in tensoes_ff.items()
                    ),
                    "Tensões Fase-Neutro": ", ".join(
                        f"{k}: {v}" for k, v in tensoes_fn.items()
                    ),
                    "Tensões Fase-Terra": ", ".join(
                        f"{k}: {v}" for k, v in tensoes_ft.items()
                    ),
                    "Tensão N/T": tensao_nt_dim,
                    "Correntes": ", ".join(f"{k}: {v}" for k, v in correntes.items()),
                    "Tipo de Cabos": st.session_state.get("tipo_cabos", tipo_cabos),
                    "Área Total Ocupada": st.session_state.get(
                        "area_total_ocupada", 0.0
                    ),
                    "Tamanho do Eletroduto": st.session_state.get(
                        "tamanho_eletroduto", ""
                    ),
                    "Disjuntor Recomendado": st.session_state.get(
                        "disjuntor_recomendado", ""
                    ),
                    "IDR Recomendado": st.session_state.get("idr_recomendado", ""),
                    "DPS Recomendado": st.session_state.get("dps_recomendado", ""),
                    "Barra Pente Recomendada": st.session_state.get(
                        "barra_pente_recomendado", ""
                    ),
                    "Tipo de Quadro Selecionado": st.session_state.get(
                        "tipo_quadro_resumo", ""
                    ),
                    "Quadro PVC Recomendado": st.session_state.get(
                        "quadro_pvc_recomendado", ""
                    ),
                    "Quadro Metálico Recomendado": st.session_state.get(
                        "quadro_metalico_recomendado", ""
                    ),
                    "Disjuntor Caixa Moldada": st.session_state.get(
                        "disjuntor_caixa_moldada", ""
                    ),
                    "Modelo Disjuntor Caixa Moldada": st.session_state.get(
                        "modelo_disjuntor_caixa_moldada", ""
                    ),
                    "Barra Roscada": st.session_state.get("barra_roscada", ""),
                    "Material Barra Roscada": st.session_state.get(
                        "barra_roscada_material", ""
                    ),
                }

                df_dimensionamento = pd.DataFrame([dados_dimensionamento])

                if filepath.exists():
                    if load_workbook is None:
                        st.error(
                            "Não foi possível atualizar o arquivo Excel porque a biblioteca "
                            "'openpyxl' não está instalada. Execute `pip install openpyxl` e tente novamente."
                        )
                    else:
                        book = load_workbook(filepath)
                        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                            writer.book = book
                            writer.sheets = {ws.title: ws for ws in book.worksheets}
                            startrow = book.active.max_row
                            df_dimensionamento.to_excel(
                                writer, index=False, header=False, startrow=startrow
                            )
                        st.success(f"Dados salvos em {filepath}")
                else:
                    try:
                        df_dimensionamento.to_excel(filepath, index=False)
                    except ImportError:
                        st.error(
                            "Não foi possível criar o arquivo Excel porque a biblioteca "
                            "'openpyxl' não está instalada. Execute `pip install openpyxl` e tente novamente."
                        )
                    else:
                        st.success(f"Dados salvos em {filepath}")
