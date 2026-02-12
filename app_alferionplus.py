import streamlit as st
from datetime import datetime
import pandas as pd

def inicializa_session_state():
    defaults = {
        "ordem_venda": "",
        "cliente": "",
        "endereco": "",
        "email": "",
        "tipo_servico": "",
        "tipo_local": "",
        "tecnico": "",
        "data_hora": datetime.today(),
        "deslocamento_necessario": "Não",
        "distancia_km": 0.0,
        "tempo_viagem": "",
        "custo_pedagios": 0.0,
        "tensao_rs": "",
        "tensao_rt": "",
        "tensao_st": "",
        "tensao_rn": "",
        "tensao_sn": "",
        "tensao_tn": "",
        "tensao_rtt": "",
        "tensao_stt": "",
        "tensao_ttt": "",
        "tensao_n_t": "",
        "corrente_r": "",
        "corrente_s": "",
        "corrente_t": "",
        "direcao": "↑",
        "trecho": 0.0,
        "direcao_quadro": "↑",
        "trecho_quadro": 0.0,
        "possui_carregador": "",
        "quantidade_carregadores": 1,
        "quadro_distribuicao": "",
        "potencia_carregador": "",
        "pot_outro_valor": 0.0,
        "marca_carregadores": "",
        "tipo_conectividade": "",
        "monofasica": False,
        "bifasica": False,
        "trifasica": False,
        "dj_disjuntor": False,
        "dj_fusivel": False,
        "dj_outro": False,
        "corrente_disjuntor": "",
        "bitola_cabos": "",
        "sistema_aterramento": "TN-S",
        "barra_neutro_terra": "",
        "espaco_dj_saida": "",
        "tem_medidor": "Não",
        "medidor": "",
        "barra_roscada": "Não",
        "barra_roscada_material": "",
        "tem_tomada_industrial": "Não",
        "tomada_industrial": "",
        "disjuntor_caixa_moldada": "Não",
        "modelo_disjuntor_caixa_moldada": "",
        "custos_modelo_disjuntor_caixa_moldada": "",
        "dimensionamento_modelo_disjuntor_caixa_moldada": "",
        "eletrocalha": "Não",
        "metros_eletrocalha": 0.0,
        "dimensoes_eletrocalha": "",
        "obra_civil": "Não",
        "infra_rede": "Não",
        "andaime": "Não",
        "transformador": "Não",
        "totem": "Não",
        "pintura_vaga": "Não",
        "pintura_eletrodutos": "Não",
        "caminhao_munk": "Não",
        "projeto_unifilar": "Não",
        "planta_baixa": "Não",
        "sem_escolha_vaga": "Não",
        "observacoes": "",
        "recados": "",
        "recado_projeto_unifilar": "",
        "recado_planta_baixa": "",
        "recado_sem_escolha_vaga": "",
        # Percursos é lista, mas inicialize fora (não com defaults)
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    # Lista de percursos, especial
    if "percursos" not in st.session_state:
        st.session_state["percursos"] = []
    if "percursos_quadro" not in st.session_state:
        st.session_state["percursos_quadro"] = []

inicializa_session_state()


from Deslocamento import (
    tempo_para_minutos,
    calcula_custo_deslocamento,
)
from dimensionamento import render_dimensionamento_tab
from calculo_servico import render_calculo_servico_tab
from custos import format_currency, render_custos_tab
from orcamento import render_orcamento_tab

st.set_page_config(page_title="Formulário de Visita Técnica", layout="centered")

# Mantém as abas visíveis ao rolar a página
# Usa seletores compatíveis com diferentes versões do Streamlit
st.markdown(
    """
    <style>
    .stTabs [role="tablist"],
    .stTabs [data-baseweb="tab-list"] {
        position: sticky;
        top: 0;
        z-index: 999;
        background-color: white;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

tab_visita, tab_dimensionamento, tab_custos, tab_calculo_servico, tab_orcamento, tab_recados = st.tabs([
    "Visita", "Dimensionamento", "Custos", "Cálculo de serviço", "Orçamento", "Recados"
])

# Inicializa config de deslocamento com padrões caso não exista
if "desloc_config" not in st.session_state:
    st.session_state["desloc_config"] = {
        "valor_combustivel": 7.50,
        "consumo_medio": 13.0,
        "valor_por_km": 0.0,
        "adicional_noturno": 350.0,
        "outros_adicionais": 0.0,
        "valor_refeicao": 60.0,
        "valor_tecnico": 100.0,
        "margem_percentual": 30.0,
    }

with tab_visita:
    st.title("📊 Formulário de Visita Técnica")

    with st.expander("🔹 Dados da Visita", expanded=True):
        col_a, col_b = st.columns([1, 1])
        with col_a:
            ordem_venda = st.text_input("Ordem de Venda", key="ordem_venda")
        with col_b:
            cliente = st.text_input("Nome do Cliente", key="cliente")

        col_c, col_d = st.columns([1, 1])
        with col_c:
            endereco = st.text_input("Endereço da Instalação", key="endereco")
        with col_d:
            email = st.text_input("Email para Contato", key="email")

        col1, col2 = st.columns([1, 1])
        with col1:
            tipo_servico = st.selectbox(
                "Tipo de Serviço",
                [
                    "",
                    "Instalação",
                    "Manutenção",
                    "Manutenção Preventiva",
                    "Manutenção Corretiva",
                    "Análise de Energia",
                    "Desinstalação",
                ],
                key="tipo_servico",
            )

        with col2:
            tipo_local = st.selectbox(
                "Tipo de Local",
                [
                    "",
                    "Outro",
                    "Comércio",
                    "Condomínio",
                    "Construção",
                    "Eletroposto",
                    "Empresa",
                    "Estacionamento",
                    "Mercado",
                    "Residencial",
                    "Shopping",
                    "Centro de Logística",
                ],
                key="tipo_local",
            )

        col3, col4 = st.columns([1, 1])
        with col3:
            tecnico = st.text_input("Técnico Responsável pela Visita", key="tecnico")
        with col4:
            data_hora = st.date_input("Data da Visita", key="data_hora")

        col5, col6 = st.columns([1.2, 1])
        with col5:
            cpf_cnpj = st.text_input("CPF / CNPJ", key="cpf_cnpj")

        with col6:
            deslocamento_necessario = st.radio(
                "Deslocamento?",
                ["Sim", "Não"],
                horizontal=True,
                key="deslocamento_necessario",
            )

        if deslocamento_necessario == "Sim":
            col7, col8, col9 = st.columns([1, 1, 1])
            with col7:
                distancia_km = st.number_input(
                    "Distância (km)",
                    min_value=0.0,
                    step=1.0,
                    key="distancia_km",
                )
            with col8:
                tempo_viagem = st.text_input(
                    "Tempo de Viagem (ex: 1h20min)",
                    key="tempo_viagem",
                )
            with col9:
                custo_pedagios = st.number_input(
                    "Custo com Pedágios (R$)",
                    min_value=0.0,
                    step=1.0,
                    key="custo_pedagios",
                )

            minutos = tempo_para_minutos(tempo_viagem)
            custo_total = calcula_custo_deslocamento(
                distancia_km,
                tempo_viagem,
                custo_pedagios,
                st.session_state["desloc_config"],
            )
            st.session_state["total_custo_deslocamento"] = custo_total
            st.success(f"💰 Custo estimado de deslocamento: R$ {custo_total:,.2f}")

        else:
            distancia_km = 0
            tempo_viagem = ""
            custo_pedagios = 0
            st.session_state["total_custo_deslocamento"] = 0.0

    with st.expander("🚗 Informações do Carregador", expanded=False):
        col_possui, col_quantidade = st.columns([1, 1])
        with col_possui:
            possui_carregador = st.radio(
                "Cliente já possui carregador?",
                ["", "Sim", "Não"],
                horizontal=True,
                key="possui_carregador",
            )

        with col_quantidade:
            quantidade_carregadores = st.number_input(
                "Quantidade de Carregadores",
                min_value=0,
                step=1,
                key="quantidade_carregadores",
            )

        quantidade_carregadores_int = int(quantidade_carregadores)

        if quantidade_carregadores_int > 1:
            st.radio(
                "Quadro de distribuição",
                ["", "Sim", "Não"],
                horizontal=True,
                key="quadro_distribuicao",
            )
        else:
            st.session_state["quadro_distribuicao"] = ""

        if quantidade_carregadores_int <= 1:
            st.markdown("**Potência dos carregadores:**")
        else:
            st.markdown("**Potências dos carregadores:**")

        opcoes_potencia = [
            "",
            "0 kW",
            "1,9 kW",
            "3,7 kW",
            "7,4 kW",
            "11 kW",
            "22 kW",
            "44 kW",
            "Outro",
        ]

        if quantidade_carregadores_int <= 1:
            potencia_selecionada = st.radio(
                "Selecione a potência",
                opcoes_potencia,
                horizontal=True,
                key="potencia_carregador_1",
            )
            if potencia_selecionada == "Outro":
                st.number_input(
                    "Potência personalizada do carregador (kW)",
                    min_value=0.0,
                    step=0.1,
                    key="pot_outro_valor_1",
                )
        else:
            for i in range(1, quantidade_carregadores_int + 1):
                potencia_selecionada = st.radio(
                    f"Potência do carregador {i}",
                    opcoes_potencia,
                    horizontal=True,
                    key=f"potencia_carregador_{i}",
                )
                if potencia_selecionada == "Outro":
                    st.number_input(
                        f"Potência personalizada do carregador {i} (kW)",
                        min_value=0.0,
                        step=0.1,
                        key=f"pot_outro_valor_{i}",
                    )

        col_marca, col_tipo = st.columns([1, 1])
        with col_marca:
            marca_carregadores = st.selectbox(
                "Marca dos Carregadores",
                [
                    "",
                    "Outro",
                    "Schneider",
                    "ABB",
                    "E-Wolf",
                    "Efacec",
                    "WEG",
                    "BWM",
                    "Neocharge",
                    "Incharge",
                    "Intelbras",
                    "Zletric",
                ],
                key="marca_carregadores",
            )

        with col_tipo:
            tipo_conectividade = st.radio(
                "Standard ou Smart",
                ["", "Básico", "Smart"],
                horizontal=True,
                key="tipo_conectividade",
            )

    with st.expander("📊 Dados Visuais", expanded=False):
        st.write("Alimentação elétrica")
        alimentacao = st.radio(
            "Selecione o tipo de alimentação elétrica",
            ["", "Monofásica", "Bifásica", "Trifásica"],
            horizontal=True,
            key="alimentacao",
        )

        col_dj, col_corrente = st.columns([1, 1])
        with col_dj:
            st.write("DJ (entrada)")
            dj_disjuntor = st.checkbox("Disjuntor", key="dj_disjuntor")
            dj_fusivel = st.checkbox("Fusível", key="dj_fusivel")
            dj_outro = st.checkbox("Outro", key="dj_outro")
        with col_corrente:
            corrente_disjuntor = st.selectbox(
                "Corrente de Desarme do DJ (A)",
                [
                    "",
                    "abaixo de 40A", "50A", "63A", "70A", "80A", "90A", "100A", "125A", "150A", "175A",
                    "200A", "250A", "300A", "350A", "400A", "450A", "500A", "600A", "800A", "1000A"
                ], key="corrente_disjuntor"
            )

        col_bitola, col_sistema = st.columns([1, 1])
        with col_bitola:
            bitola_cabos = st.selectbox(
                "Bitola dos cabos de entrada",
                [
                    "",
                    "abaixo de 6mm", "6mm", "10mm", "16mm", "25mm", "35mm", "50mm", "70mm", "95mm",
                    "120mm", "150mm", "185mm", "240mm", "300mm", "400mm", "500mm",
                    "2x90mm", "2x120mm", "2x150mm", "2x185mm", "2x240mm", "3x240mm"
                ], key="bitola_cabos"
            )

        with col_sistema:
            sistema_aterramento = st.selectbox(
                "Sistema de aterramento",
                [
                    "",
                    "Sem aterramento", "TN-S", "TN-C", "TN-C-S", "TT", "IT"
                ], key="sistema_aterramento"
            )

        col_barra, col_espaco = st.columns([1, 1])
        with col_barra:
            barra_neutro_terra = st.selectbox(
                "Barra de neutro e terra",
                [
                    "",
                    "Com Barramento N/T",
                    "Sem Barramento",
                    "Barramento de Neutro",
                    "Barramento de Terra",
                ],
                key="barra_neutro_terra",
            )

        with col_espaco:
            espaco_dj_saida = st.selectbox(
                "Espaço DJ Saída",
                ["", "Com Espaço DJ", "Sem Espaço DJ", "Adaptar Espaço DJ"],
                key="espaco_dj_saida",
            )

    with st.expander("🔎 Medições", expanded=False):
        with st.container():
            st.markdown(
                "<div style='border:2px solid red;padding:10px;'>Tensão entre fases:",
                unsafe_allow_html=True,
            )
            col_rs, col_rt, col_st = st.columns(3)
            with col_rs:
                tensao_rs = st.text_input("R/S", key="tensao_rs")
            with col_rt:
                tensao_rt = st.text_input("R/T", key="tensao_rt")
            with col_st:
                tensao_st = st.text_input("S/T", key="tensao_st")

        with st.container():
            st.markdown(
                "<div style='border:2px solid lightblue;padding:10px;'>Tensão entre fases e neutro:",
                unsafe_allow_html=True,
            )
            col_rn, col_sn, col_tn = st.columns(3)
            with col_rn:
                tensao_rn = st.text_input("R/N", key="tensao_rn")
            with col_sn:
                tensao_sn = st.text_input("S/N", key="tensao_sn")
            with col_tn:
                tensao_tn = st.text_input("T/N", key="tensao_tn")

        with st.container():
            st.markdown(
                "<div style='border:2px solid green;padding:10px;'>Tensão entre fases e terra:",
                unsafe_allow_html=True,
            )
            col_rtt, col_stt, col_ttt = st.columns(3)
            with col_rtt:
                tensao_rtt = st.text_input("R/T Terra", key="tensao_rtt")
            with col_stt:
                tensao_stt = st.text_input("S/T Terra", key="tensao_stt")
            with col_ttt:
                tensao_ttt = st.text_input("T/T Terra", key="tensao_ttt")

        n_t = st.text_input("N/T", key="tensao_n_t")

        with st.container():
            st.markdown(
                "<div style='border:2px solid black;padding:10px;'>Corrente registrada:",
                unsafe_allow_html=True,
            )
            col_r, col_s, col_t = st.columns(3)
            with col_r:
                corrente_r = st.text_input("R", key="corrente_r")
            with col_s:
                corrente_s = st.text_input("S", key="corrente_s")
            with col_t:
                corrente_t = st.text_input("T", key="corrente_t")


    total_quadro = 0.0
    if st.session_state.get("quadro_distribuicao", "") == "Sim":
        with st.expander("📊 Soma da Distância do Quadro de Distribuição com Direções", expanded=False):
            if "percursos_quadro" not in st.session_state:
                st.session_state.percursos_quadro = []

            with st.form(key="percurso_quadro_form"):
                col_a, col_b = st.columns([2, 1])
                with col_a:
                    direcao_quadro = st.selectbox(
                        "Direção", ["↑", "↓", "→", "←", "↷", "↶"], key="direcao_quadro"
                    )
                with col_b:
                    trecho_quadro = st.number_input(
                        "Distância (m)", min_value=0.0, step=0.5, key="trecho_quadro"
                    )
                submitted_quadro = st.form_submit_button("Adicionar trecho")

            if submitted_quadro:
                st.session_state.percursos_quadro.append(
                    (st.session_state.direcao_quadro, st.session_state.trecho_quadro)
                )

            if st.session_state.percursos_quadro:
                total_quadro = sum(t[1] for t in st.session_state.percursos_quadro)
                total_quadro_str = f"{total_quadro:g}"
                st.markdown("**Trechos registrados:**")
                for i, (d, t) in enumerate(st.session_state.percursos_quadro):
                    col1, col2 = st.columns([4, 1])
                    with col1:
                        st.markdown(f"{i+1}. {d} {t} m")
                    with col2:
                        if st.button("❌", key=f"delete_quadro_{i}"):
                            st.session_state.percursos_quadro.pop(i)
                            st.rerun()
                st.markdown(f"**Total: {total_quadro_str} m**")

    st.session_state["distancia_alimentacao_distribuicao"] = total_quadro
    with st.expander("🛠️ Informações Técnicas", expanded=False):
        st.text_input(
            "Distãncia entre Alimentação e Distribuição",
            value=f"{total_quadro:g}",
            disabled=True,
        )

    with st.expander("📊 Soma da Distância com Direções", expanded=False):
        if "percursos" not in st.session_state:
            st.session_state.percursos = []

        with st.form(key="percurso_form"):
            col_a, col_b = st.columns([2, 1])
            with col_a:
                direcao = st.selectbox(
                    "Direção", ["↑", "↓", "→", "←", "↷", "↶"], key="direcao"
                )
            with col_b:
                trecho = st.number_input(
                    "Distância (m)", min_value=0.0, step=0.5, key="trecho"
                )
            submitted = st.form_submit_button("Adicionar trecho")

        if submitted:
            st.session_state.percursos.append(
                (st.session_state.direcao, st.session_state.trecho)
            )

        if st.session_state.percursos:
            total = sum(t[1] for t in st.session_state.percursos)
            total_str = f"{total:g}"
            st.markdown("**Trechos registrados:**")
            for i, (d, t) in enumerate(st.session_state.percursos):
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.markdown(f"{i+1}. {d} {t} m")
                with col2:
                    if st.button(f"❌", key=f"delete_{i}"):
                        st.session_state.percursos.pop(i)
                        st.rerun()
            st.markdown(f"**Total: {total_str} m**")

    with st.expander("📦 Material Adicional", expanded=False):
        col_dcm, col_barra, col_eletro = st.columns([1, 1, 1])
        with col_dcm:
            disjuntor_caixa_moldada = st.radio(
                "Disjuntor Caixa Moldada", ["Não", "Sim"],
                horizontal=True,
                key="disjuntor_caixa_moldada",
            )
        with col_barra:
            barra_roscada = st.radio(
                "Barra Roscada (vergalhão)", ["Não", "Sim"],
                horizontal=True,
                key="barra_roscada",
            )
        with col_eletro:
            eletrocalha = st.radio(
                "Eletrocalha", ["Não", "Sim"],
                horizontal=True,
                key="eletrocalha",
            )
            if eletrocalha == "Sim":
                metros_eletrocalha = st.number_input(
                    "Quantos metros?", min_value=0.0, step=1.0,
                    key="metros_eletrocalha",
                )
            else:
                metros_eletrocalha = 0
                st.session_state["dimensoes_eletrocalha"] = ""

        col_tomada, col_medidor, col_transformador, col_totem = st.columns(4)
        with col_tomada:
            tomada_industrial_flag = st.radio(
                "Tomada industrial",
                ["Não", "Sim"],
                horizontal=True,
                key="tem_tomada_industrial",
            )
            if tomada_industrial_flag != "Sim":
                st.session_state["tomada_industrial"] = ""
        with col_medidor:
            medidor_flag = st.radio(
                "Medidor",
                ["Não", "Sim"],
                horizontal=True,
                key="tem_medidor",
            )
            if medidor_flag != "Sim":
                st.session_state["medidor"] = ""
        with col_transformador:
            transformador = st.radio(
                "Transformador",
                ["Não", "Sim"],
                horizontal=True,
                key="transformador",
            )
        with col_totem:
            totem = st.radio(
                "Totem",
                ["Não", "Sim"],
                horizontal=True,
                key="totem",
            )

    with st.expander("🔧 Serviços Adicionais", expanded=False):
        col_obra, col_infra, col_andaime = st.columns(3)
        with col_obra:
            obra_civil = st.radio(
                "Obra Civil", ["Não", "Sim"],
                horizontal=True,
                key="obra_civil"
            )
        with col_infra:
            infra_rede = st.radio(
                "Infra de rede", ["Não", "Sim"],
                horizontal=True,
                key="infra_rede"
            )
        with col_andaime:
            andaime = st.radio(
                "Andaime", ["Não", "Sim"],
                horizontal=True,
                key="andaime"
            )
        col_pintura_vaga, col_pintura_eletrodutos, col_caminhao = st.columns(3)
        with col_pintura_vaga:
            pintura_vaga = st.radio(
                "Pintura da vaga", ["Não", "Sim"],
                horizontal=True,
                key="pintura_vaga"
            )
        with col_pintura_eletrodutos:
            pintura_eletrodutos = st.radio(
                "Pintura do eletrodutos", ["Não", "Sim"],
                horizontal=True,
                key="pintura_eletrodutos"
            )
        with col_caminhao:
            caminhao_munk = st.radio(
                "Caminhão Munk", ["Não", "Sim"],
                horizontal=True,
                key="caminhao_munk"
            )

    with st.expander("📌 Pendências", expanded=False):
        pend_col1, pend_col2, pend_col3 = st.columns(3)
        with pend_col1:
            projeto_unifilar = st.radio(
                "Pedir Projeto Unifilar", ["Não", "Sim"],
                horizontal=True,
                key="projeto_unifilar"
            )
        with pend_col2:
            planta_baixa = st.radio(
                "Pedir Planta Baixa das vagas", ["Não", "Sim"],
                horizontal=True,
                key="planta_baixa"
            )
        with pend_col3:
            sem_escolha_vaga = st.radio(
                "Cliente não escolheu as vagas", ["Não", "Sim"],
                horizontal=True,
                key="sem_escolha_vaga"
            )

    observacoes = st.text_area("Observações", key="observacoes")

    if st.button("Salvar Dados da Visita"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            load_workbook = None
        from pathlib import Path
        # Recalcula soma de percursos
        total_dist = sum(t[1] for t in st.session_state.percursos) if "percursos" in st.session_state else 0
        total_dist_quadro = (
            sum(t[1] for t in st.session_state.percursos_quadro)
            if "percursos_quadro" in st.session_state
            else 0
        )
        quantidade_carregadores_int = int(st.session_state.get("quantidade_carregadores", 0))
        potencias_carregadores = []
        for i in range(1, quantidade_carregadores_int + 1):
            potencia_carregador = st.session_state.get(f"potencia_carregador_{i}", "")
            if potencia_carregador == "Outro":
                potencia_carregador_val = st.session_state.get(f"pot_outro_valor_{i}", 0.0)
                potencia_carregador_val = f"{potencia_carregador_val} kW"
            else:
                potencia_carregador_val = potencia_carregador
            potencias_carregadores.append(potencia_carregador_val)

        if quantidade_carregadores_int <= 1:
            potencia_carregador_val = potencias_carregadores[0] if potencias_carregadores else ""
        else:
            potencia_carregador_val = "; ".join(
                [f"Carregador {idx + 1}: {pot}" for idx, pot in enumerate(potencias_carregadores)]
            )
        # Define o tipo de alimentação elétrica selecionado
        alimentacao = st.session_state.get("alimentacao", "")
        monofasica = alimentacao == "Monofásica"
        bifasica = alimentacao == "Bifásica"
        trifasica = alimentacao == "Trifásica"
        # Coleta dados do formulário em um dicionário
        tomada_industrial_flag = st.session_state.get("tem_tomada_industrial", "Não")
        tomada_industrial_modelo = st.session_state.get("tomada_industrial", "")
        medidor_flag = st.session_state.get("tem_medidor", "Não")
        medidor_modelo = st.session_state.get("medidor", "")

        data = {
            "Ordem de Venda": ordem_venda,
            "Cliente": cliente,
            "CPF / CNPJ": cpf_cnpj,
            "Endereço da Instalação": endereco,
            "Email": email,
            "Tipo de Serviço": tipo_servico,
            "Tipo de Local": tipo_local,
            "Técnico Responsável": tecnico,
            "Data da Visita": data_hora,
            "Deslocamento": deslocamento_necessario,
            "Distância (km)": distancia_km,
            "Tempo de Viagem": tempo_viagem,
            "Custo com Pedágios (R$)": custo_pedagios,
            "R/S": tensao_rs,
            "R/T": tensao_rt,
            "S/T": tensao_st,
            "R/N": tensao_rn,
            "S/N": tensao_sn,
            "T/N": tensao_tn,
            "R/T Terra": tensao_rtt,
            "S/T Terra": tensao_stt,
            "T/T Terra": tensao_ttt,
            "N/T": n_t,
            "Corrente R": corrente_r,
            "Corrente S": corrente_s,
            "Corrente T": corrente_t,
            "Corrente Calculada": st.session_state.get("corrente_calculada", ""),
            "Soma Distância (m)": total_dist,
            "Soma Distância Quadro de Distribuição (m)": total_dist_quadro,
            "Cliente já possui carregador": possui_carregador,
            "Quantidade Carregadores": quantidade_carregadores,
            "Quadro de distribuição": st.session_state.get("quadro_distribuicao", ""),
            "Potência Carregador": potencia_carregador_val,
            "Marca Carregadores": marca_carregadores,
            "Standard ou Smart": tipo_conectividade,
            "Monofásica": monofasica,
            "Bifásica": bifasica,
            "Trifásica": trifasica,
            "DJ Disjuntor": dj_disjuntor,
            "DJ Fusível": dj_fusivel,
            "DJ Outro": dj_outro,
            "Corrente Desarme DJ (A)": corrente_disjuntor,
            "Bitola Cabos": bitola_cabos,
            "Sistema Aterramento": sistema_aterramento,
            "Barra Neutro e Terra": barra_neutro_terra,
            "Espaço DJ Saída": espaco_dj_saida,
            "Medidor": medidor_flag,
            "Modelo Medidor": medidor_modelo,
            "Barra Roscada": barra_roscada,
            "Modelo Barra Roscada (vergalhão)": st.session_state.get(
                "barra_roscada_material", ""
            ),
            "Tomada Industrial": tomada_industrial_flag,
            "Modelo Tomada Industrial": tomada_industrial_modelo,
            "Disjuntor Caixa Moldada": disjuntor_caixa_moldada,
            "Modelo Disjuntor Caixa Moldada": st.session_state.get(
                "modelo_disjuntor_caixa_moldada", ""
            ),
            "Eletrocalha": eletrocalha,
            "Metros Eletrocalha": metros_eletrocalha,
            "Obra Civil": obra_civil,
            "Infra de rede": infra_rede,
            "Andaime": andaime,
            "Transformador": transformador,
            "Totem": totem,
            "Pintura da vaga": pintura_vaga,
            "Pintura dos eletrodutos": pintura_eletrodutos,
            "Caminhão Munk": caminhao_munk,
            "Pedir Projeto Unifilar": projeto_unifilar,
            "Pedir Planta Baixa das vagas": planta_baixa,
            "Cliente não escolheu vagas": sem_escolha_vaga,
            "Observações": observacoes,
        }

        for idx, potencia in enumerate(potencias_carregadores, start=1):
            data[f"Potência Carregador {idx}"] = potencia

        df = pd.DataFrame([data])
        # Define pasta de saída "Docs Salvos" ao lado deste script
        docs_dir = Path(__file__).with_name("Docs Salvos")
        docs_dir.mkdir(exist_ok=True)
        # Nome do arquivo: "Dados da Visita Técnica" + Ordem de Venda
        base_name = "Dados da Visita Técnica"
        filename = f"{base_name} {ordem_venda}.xlsx" if ordem_venda else f"{base_name}.xlsx"
        filepath = docs_dir / filename
        # Se o arquivo já existe, acrescenta a nova linha
        if filepath.exists():
            if load_workbook is None:
                st.error(
                    "Não foi possível atualizar o arquivo Excel porque a biblioteca "
                    "'openpyxl' não está instalada. Execute `pip install openpyxl` e tente novamente."
                )
            else:
                book = load_workbook(filepath)
                with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
                    writer.book = book
                    writer.sheets = {ws.title: ws for ws in book.worksheets}
                    # Determina a primeira linha vazia para inserir dados
                    startrow = book.active.max_row
                    df.to_excel(writer, index=False, header=False, startrow=startrow)
                st.success(f"Dados salvos em {filepath}")
        else:
            try:
                df.to_excel(filepath, index=False)
            except ImportError:
                st.error(
                    "Não foi possível criar o arquivo Excel porque a biblioteca "
                    "'openpyxl' não está instalada. Execute `pip install openpyxl` e tente novamente."
                )
            else:
                st.success(f"Dados salvos em {filepath}")
render_dimensionamento_tab(tab_dimensionamento)
render_custos_tab(tab_custos)
render_calculo_servico_tab(tab_calculo_servico, format_currency)
render_orcamento_tab(tab_orcamento)


def render_recados_tab(tab):
    """Renderiza a aba de Recados."""

    with tab:
        st.title("📝 Recados")
        st.text_area(
            "Registre observações importantes para a equipe.",
            key="recados",
            height=200,
        )

        pendencias_campos = {
            "projeto_unifilar": "Pedir Projeto Unifilar",
            "planta_baixa": "Pedir Planta Baixa das vagas",
            "sem_escolha_vaga": "Cliente não escolheu as vagas",
        }

        for state_key, label in pendencias_campos.items():
            if st.session_state.get(state_key) == "Sim":
                st.text_area(
                    label,
                    key=f"recado_{state_key}",
                    height=120,
                )


render_recados_tab(tab_recados)
